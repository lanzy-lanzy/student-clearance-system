import json
import logging
from datetime import datetime
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.files.storage import default_storage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_POST, require_http_methods
from django.views.generic import TemplateView, ListView, DetailView

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Local imports
from core.models import (
    ClearanceRequest, Clearance, Staff, Student,
    Office, ProgramChair, User, Dean, Course,
    UserProfile, SEMESTER_CHOICES
)

# Set up logging
logger = logging.getLogger(__name__)

# Utility Functions
def get_current_school_year():
    current_date = datetime.now()
    return (f"{current_date.year - 1}-{current_date.year}"
            if current_date.month <= 5
            else f"{current_date.year}-{current_date.year + 1}")

def get_current_semester():
    current_month = datetime.now().month
    if 1 <= current_month <= 5:
        return "Second"
    elif 6 <= current_month <= 7:
        return "Summer"
    return "First"

def get_school_years():
    """Get a list of school years for the past 5 years and next 2 years"""
    current_year = timezone.now().year
    years = []

    # Past 5 years
    for i in range(5, 0, -1):
        past_year = current_year - i
        years.append(f"{past_year}-{past_year + 1}")

    # Current year
    years.append(f"{current_year}-{current_year + 1}")

    # Next 2 years
    for i in range(1, 3):
        future_year = current_year + i
        years.append(f"{future_year}-{future_year + 1}")

    return years

def is_program_chair(user):
    return hasattr(user, 'programchair')

# Authentication Views
def user_login(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'student'):
            return redirect('student_dashboard')
        elif hasattr(request.user, 'programchair'):
            return redirect('program_chair_dashboard')
        elif hasattr(request.user, 'staff'):
            # Check if the staff member is a dormitory owner
            if request.user.staff.is_dormitory_owner:
                return redirect('bh_owner_dashboard')
            else:
                return redirect('staff_dashboard')
        elif request.user.is_superuser:
            return redirect('admin_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user:
            if not user.is_active:
                messages.error(request, 'Your account is not yet approved.')
                return redirect('login')

            login(request, user)
            if hasattr(user, 'student'):
                return redirect('student_dashboard')
            elif hasattr(user, 'programchair'):
                return redirect('program_chair_dashboard')
            elif hasattr(user, 'staff'):
                # Check if the staff member is a dormitory owner
                if user.staff.is_dormitory_owner:
                    return redirect('bh_owner_dashboard')
                else:
                    return redirect('staff_dashboard')
            elif user.is_superuser:
                return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid credentials.')

    return render(request, 'registration/login.html')

def user_logout(request):
    logout(request)
    return redirect('login')

def register(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        try:
            # Debug logging
            course_id = request.POST.get('course')
            print(f"Received course ID: {course_id}")
            print(f"All POST data: {request.POST}")

            username = request.POST.get('username')
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists.')
                return redirect('register')

            email = request.POST.get('email')
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists.')
                return redirect('register')

            user = User.objects.create_user(
                username=username,
                email=email,
                password=request.POST.get('password'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                is_active=False
            )

            try:
                course = Course.objects.get(id=course_id)
                print(f"Found course: {course.code} - {course.name}")
            except Course.DoesNotExist:
                print(f"No course found with ID: {course_id}")
                available_courses = Course.objects.all()
                print(f"Available course IDs: {[c.id for c in available_courses]}")
                raise Exception(f"Course with ID {course_id} does not exist")

            # Get the program chair for the selected course's dean
            program_chair = ProgramChair.objects.filter(dean=course.dean).first()

            # Get dormitory owner if student is a boarder
            dormitory_owner = None
            if request.POST.get('is_boarder') == 'on' and request.POST.get('dormitory_owner'):
                dormitory_owner = Staff.objects.filter(id=request.POST.get('dormitory_owner')).first()

            Student.objects.create(
                user=user,
                student_id=request.POST.get('student_id'),
                course=course,
                program_chair=program_chair,
                dormitory_owner=dormitory_owner,
                year_level=request.POST.get('year_level'),
                contact_number=request.POST.get('contact_number'),
                is_boarder=request.POST.get('is_boarder') == 'on',
                is_approved=False
            )

            messages.success(request, 'Registration successful! Awaiting admin approval.')
            return redirect('login')

        except Exception as e:
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Registration error: {str(e)}')

    context = {
        'deans': Dean.objects.all().order_by('name'),
        'dormitory_owners': Staff.objects.filter(is_dormitory_owner=True)
    }
    return render(request, 'registration/register.html', context)

# Basic Views
def home(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'programchair'):
            return redirect('program_chair_dashboard')
        elif hasattr(request.user, 'student'):
            return redirect('student_dashboard')
        elif hasattr(request.user, 'staff'):
            # Check if the staff member is a dormitory owner
            if request.user.staff.is_dormitory_owner:
                return redirect('bh_owner_dashboard')
            else:
                return redirect('staff_dashboard')
        elif request.user.is_superuser:
            return redirect('admin_dashboard')
    return render(request, 'home.html')

# Student Views
@login_required
def student_dashboard(request):
    try:
        student = request.user.student
        clearances = Clearance.objects.filter(student=student).order_by('-school_year', '-semester')

        # Get the latest clearance for the student
        latest_clearance = clearances.first()

        # Get all clearance requests for the student
        if latest_clearance:
            clearance_requests = ClearanceRequest.objects.filter(
                clearance=latest_clearance
            ).select_related('office', 'reviewed_by')
        else:
            clearance_requests = []

        # Get student's course and other related information
        course_info = student.course

        return render(request, 'core/student_dashboard.html', {
            'student_info': student,
            'clearances': clearances,
            'latest_clearance': latest_clearance,
            'clearance_requests': clearance_requests,
            'course_info': course_info,
        })
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('home')

@login_required
def student_profile(request):
    try:
        student = request.user.student

        if request.method == 'POST' and request.POST.get('action') == 'change_password':
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            # Validate passwords
            if not request.user.check_password(current_password):
                messages.error(request, 'Current password is incorrect.')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match.')
            else:
                # Set new password with Django's built-in password hashing
                request.user.set_password(new_password)
                request.user.save()
                messages.success(request, 'Password changed successfully. Please log in with your new password.')
                return redirect('login')

        student_info = {
            'full_name': student.user.get_full_name(),
            'student_id': student.student_id,
            'course': student.course.name,
            'year_level': student.year_level,
            'program_chair': student.course.dean.name,
            'is_boarder': student.is_boarder
        }
        return render(request, 'core/student_profile.html', {
            'student': student,
            'student_info': student_info
        })
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('home')

@login_required
def create_clearance_requests(request):
    if not hasattr(request.user, 'student'):
        messages.error(request, "Student profile required.")
        return redirect('home')

    student = request.user.student

    # Get system settings
    from core.models import SystemSettings
    settings = SystemSettings.get_settings()

    # We'll check if clearance is active, but we'll still show the page with disabled controls
    # This allows students to see the current clearance period even if they can't submit requests
    clearance_active = settings.clearance_active

    # Log the clearance status for debugging
    logger.info(f"Clearance active status: {clearance_active}, School year: {settings.school_year}, Semester: {settings.semester}")

    if request.method == 'POST':
        # First check if clearance is active
        if not settings.clearance_active:
            messages.error(request, "Clearance requests are currently disabled by the administrator.")
            return redirect('student_dashboard')

        # Use system settings as default values
        school_year = request.POST.get('school_year', settings.school_year)
        semester = request.POST.get('semester', settings.semester)

        if not all([school_year, semester]):
            messages.error(request, "School year and semester required.")
            return redirect('create_clearance_requests')

        # Ensure the submitted school year and semester match the admin settings
        if school_year != settings.school_year or semester != settings.semester:
            messages.error(request, f"Clearance requests can only be created for the current period: {settings.get_semester_display()} {settings.school_year}")
            return redirect('create_clearance_requests')

        if Clearance.objects.filter(student=student, school_year=school_year, semester=semester).exists():
            messages.error(request, f"Clearance already exists for {school_year} {semester}.")
            return redirect('student_dashboard')

        if not student.is_approved:
            messages.error(request, "Account must be approved first.")
            return redirect('student_dashboard')

        if not student.program_chair:
            messages.error(request, "You must have an assigned program chair before creating clearance requests.")
            return redirect('student_dashboard')

        try:
            clearance = Clearance.objects.create(
                student=student,
                school_year=school_year,
                semester=semester,
                is_cleared=False
            )

            # Log student information
            logger.info(f"Creating clearance for student {student.id}")

            # Get basic offices (non-program chair offices)
            required_offices = Office.objects.filter(
                Q(office_type='OTHER') | Q(office_type=student.course.dean.name)
            )

            # If student is assigned to a dormitory owner but has been deactivated,
            # exclude the dormitory office from clearance requests
            if student.dormitory_owner and not student.is_boarder:
                dormitory_offices = Office.objects.filter(name__icontains='dormitory')
                if dormitory_offices.exists():
                    required_offices = required_offices.exclude(id__in=dormitory_offices.values_list('id', flat=True))
                    logger.info(f"Excluded dormitory offices for deactivated student {student.id}")

            # Debug logging
            logger.info(f"Initial required offices: {[o.name for o in required_offices]}")
            logger.info(f"Student is_boarder: {student.is_boarder}")
            logger.info(f"Student dormitory_owner: {student.dormitory_owner}")

            # Exclude DORMITORY office for non-boarders
            if not student.is_boarder:
                required_offices = required_offices.exclude(name="DORMITORY")
                logger.info(f"Excluded DORMITORY office for non-boarder student {student.id}")

            # Exclude all offices that have program chairs assigned to them
            # First, get all program chair users
            program_chair_users = ProgramChair.objects.all().values_list('user_id', flat=True)

            # Then, find all offices that have staff with these users
            program_chair_offices = Office.objects.filter(staff__user_id__in=program_chair_users)

            # Also exclude offices with 'PROGRAM CHAIR' in the name
            program_chair_name_offices = Office.objects.filter(name__icontains="PROGRAM CHAIR")

            # Combine the office IDs to exclude
            offices_to_exclude = list(program_chair_offices.values_list('id', flat=True)) + list(program_chair_name_offices.values_list('id', flat=True))

            if offices_to_exclude:
                required_offices = required_offices.exclude(id__in=offices_to_exclude)
                logger.info(f"Excluded all offices with program chairs assigned to them")

            # Log the final list of offices
            office_names = [office.name for office in required_offices]
            logger.info(f"Final list of offices for clearance: {office_names}")

            for office in required_offices:
                ClearanceRequest.objects.create(
                    student=student,
                    clearance=clearance,
                    office=office,
                    school_year=school_year,
                    semester=semester,
                    status='pending'
                )

            messages.success(request, f"Clearance requests created for {school_year} {semester}.")
            return redirect('view_clearance_details', clearance_id=clearance.id)

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            logger.error(f"Error creating clearance requests: {str(e)}", exc_info=True)
            return redirect('student_dashboard')

    return render(request, 'core/create_clearance_requests.html', {
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
        'student': student,
        'settings': settings,
        'current_school_year': settings.school_year,
        'current_semester': settings.semester
    })

@login_required
@require_POST
def request_again(request, request_id):
    clearance_request = get_object_or_404(ClearanceRequest, id=request_id)

    if clearance_request.student.user != request.user:
        messages.error(request, "Permission denied.")
        return redirect('student_dashboard')

    if clearance_request.status != 'denied':
        messages.error(request, "Can only request again for denied clearances.")
        return redirect('view_clearance_details', clearance_id=clearance_request.clearance.id)

    try:
        clearance_request.status = 'pending'
        clearance_request.reviewed_date = None
        clearance_request.reviewed_by = None
        clearance_request.notes = None
        clearance_request.request_date = timezone.now()
        clearance_request.save()

        messages.success(request, f"Resubmitted clearance request for {clearance_request.office.name}")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('view_clearance_details', clearance_id=clearance_request.clearance.id)

@login_required
def student_clearance_history(request):
    try:
        student = request.user.student

        # Get all clearance requests for the student
        clearance_requests = ClearanceRequest.objects.filter(
            student=student
        ).select_related('office', 'reviewed_by', 'clearance').order_by('-request_date')

        # Get counts for each status before applying filters
        approved_count = clearance_requests.filter(status='approved').count()
        pending_count = clearance_requests.filter(status='pending').count()
        denied_count = clearance_requests.filter(status='denied').count()

        # Apply filters if provided
        status = request.GET.get('status', '')
        if status:
            clearance_requests = clearance_requests.filter(status=status)

        school_year = request.GET.get('school_year', '')
        if school_year:
            clearance_requests = clearance_requests.filter(school_year=school_year)

        semester = request.GET.get('semester', '')
        if semester:
            clearance_requests = clearance_requests.filter(semester=semester)

        office = request.GET.get('office', '')
        if office:
            clearance_requests = clearance_requests.filter(office__name__icontains=office)

        # Pagination
        paginator = Paginator(clearance_requests, 10)  # Show 10 requests per page
        page = request.GET.get('page')
        try:
            clearance_requests_page = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page
            clearance_requests_page = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver last page of results
            clearance_requests_page = paginator.page(paginator.num_pages)

        # Get all offices for filter dropdown
        offices = Office.objects.filter(clearance_requests__student=student).distinct()

        return render(request, 'core/student_clearance_history.html', {
            'student': student,
            'clearance_requests': clearance_requests_page,
            'school_years': get_school_years(),
            'semesters': SEMESTER_CHOICES,
            'offices': offices,
            'approved_count': approved_count,
            'pending_count': pending_count,
            'denied_count': denied_count,
            'current_filters': {
                'status': status,
                'school_year': school_year,
                'semester': semester,
                'office': office
            }
        })
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('home')

# Program Chair Views
@login_required
@user_passes_test(is_program_chair)
def program_chair_dashboard(request):
    program_chair = request.user.programchair
    view = request.GET.get('view', '')

    # Get current school year and semester
    current_year = timezone.now().year
    selected_school_year = request.GET.get('school_year', f"{current_year}-{current_year + 1}")

    # Determine current semester based on month
    month = timezone.now().month
    current_semester = "1ST"
    if 1 <= month <= 5:  # January to May
        current_semester = "2ND"
    elif 6 <= month <= 10:  # June to October
        current_semester = "1ST"
    else:  # November to December
        current_semester = "SUM"

    selected_semester = request.GET.get('semester', current_semester)

    # Generate available years (current year and 4 years back)
    available_years = [f"{year}-{year + 1}" for year in range(current_year - 4, current_year + 1)]

    # Get all students under this program chair's dean
    students = Student.objects.filter(course__dean=program_chair.dean)

    # Basic statistics for dashboard
    total_students = students.count()

    # Get clearances for the selected school year and semester
    clearances = Clearance.objects.filter(
        student__in=students,
        school_year=selected_school_year,
        semester=selected_semester
    )

    # Calculate real data counts for dashboard cards
    cleared_students = clearances.filter(is_cleared=True).count()
    pending_clearances = clearances.filter(is_cleared=False).count()

    # Handle POST requests for batch approvals
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'batch_approve':
            selected_students = request.POST.getlist('selected_students')
            approval_type = request.POST.get('approval_type')
            comments = request.POST.get('comments', '')

            if not selected_students:
                messages.error(request, "No students selected for batch approval.")
                return redirect(f"?view={view}&school_year={selected_school_year}&semester={selected_semester}")

            # Count for success message
            approved_count = 0

            # Get the batch of students
            batch_students = Student.objects.filter(id__in=selected_students)

            for student in batch_students:
                clearance = Clearance.objects.filter(
                    student=student,
                    school_year=selected_school_year,
                    semester=selected_semester
                ).first()

                if not clearance:
                    continue  # Skip if no clearance exists

                if approval_type == 'final':
                    # For final approval, we only need to mark the clearance as program_chair_approved
                    # if it's already cleared by all offices
                    if clearance.is_cleared:
                        clearance.program_chair_approved = True
                        clearance.save()
                        approved_count += 1

                elif approval_type == 'permit':
                    # For permit issuance, we need to ensure the clearance is cleared and approved
                    if clearance.is_cleared:
                        # Mark as program chair approved (which indicates permit issuance)
                        clearance.program_chair_approved = True
                        clearance.save()

                        # Here you would create a Permit record if you have a Permit model
                        # For now, we'll just count it as approved
                        approved_count += 1

            # Create a record of this batch approval (in a real implementation)
            # BatchApproval.objects.create(
            #     approved_by=request.user,
            #     approval_type=approval_type,
            #     school_year=selected_school_year,
            #     semester=selected_semester,
            #     comments=comments,
            #     student_count=approved_count
            # )

            if approved_count > 0:
                messages.success(request, f"Successfully approved {approved_count} students.")
            else:
                messages.warning(request, "No eligible students were found for approval.")

            return redirect(f"?view={view}&school_year={selected_school_year}&semester={selected_semester}")

        elif action == 'batch_print_permits':
            # Handle batch permit printing
            course_id = request.POST.get('course')
            department_id = request.POST.get('department')
            year_level = request.POST.get('year_level')
            status = request.POST.get('status')

            # Filter students based on criteria
            permit_students = students
            if course_id:
                permit_students = permit_students.filter(course_id=course_id)
            if department_id:
                permit_students = permit_students.filter(course__dean_id=department_id)
            if year_level:
                permit_students = permit_students.filter(year_level=year_level)

            # Further filter based on clearance status
            eligible_students = []
            for student in permit_students:
                clearance = Clearance.objects.filter(
                    student=student,
                    school_year=selected_school_year,
                    semester=selected_semester,
                    is_cleared=True
                ).first()

                if clearance:
                    eligible_students.append(student.id)

            if eligible_students:
                # Store eligible students in session for batch printing
                request.session['batch_print_students'] = eligible_students
                return redirect('batch_print_permits')
            else:
                messages.error(request, "No eligible students found for batch permit printing.")

    # Common context data for all views
    context = {
        'program_chair': program_chair,
        'total_students': total_students,
        'cleared_students': cleared_students,
        'pending_clearances': pending_clearances,
        'selected_school_year': selected_school_year,
        'selected_semester': selected_semester,
        'available_years': available_years,
    }

    # Handle different views
    if view == 'clearance_management':
        status_filter = request.GET.get('status', '')
        search_query = request.GET.get('search', '')

        # Calculate statistics for clearance management
        total_clearances = clearances.count() or 1  # Avoid division by zero
        cleared_count = clearances.filter(is_cleared=True).count()
        pending_count = clearances.filter(is_cleared=False).count()
        not_started_count = total_students - (cleared_count + pending_count)

        cleared_percentage = int((cleared_count / total_clearances) * 100) if total_clearances > 0 else 0
        pending_percentage = int((pending_count / total_clearances) * 100) if total_clearances > 0 else 0
        not_started_percentage = int((not_started_count / total_students) * 100) if total_students > 0 else 0

        # Filter students based on clearance status
        filtered_students = students
        if status_filter == 'cleared':
            student_ids = clearances.filter(is_cleared=True).values_list('student_id', flat=True)
            filtered_students = filtered_students.filter(id__in=student_ids)
        elif status_filter == 'pending':
            student_ids = clearances.filter(is_cleared=False).values_list('student_id', flat=True)
            filtered_students = filtered_students.filter(id__in=student_ids)
        elif status_filter == 'not_started':
            student_ids = clearances.values_list('student_id', flat=True)
            filtered_students = filtered_students.exclude(id__in=student_ids)

        # Apply search filter if provided
        if search_query:
            filtered_students = filtered_students.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(student_id__icontains=search_query)
            )

        # Paginate the results
        paginator = Paginator(filtered_students, 10)
        page = request.GET.get('page')
        students_page = paginator.get_page(page)

        # Add clearance management specific context
        context.update({
            'students': students_page,
            'cleared_count': cleared_count,
            'pending_count': pending_count,
            'not_started_count': not_started_count,
            'cleared_percentage': cleared_percentage,
            'pending_percentage': pending_percentage,
            'not_started_percentage': not_started_percentage,
        })

    elif view == 'batch_approval':
        # Get students eligible for batch approval
        # Filter by year level if provided
        year_level = request.GET.get('year_level', '')
        filtered_students = students
        if year_level:
            filtered_students = filtered_students.filter(year_level=year_level)

        # Get all students with clearances for the selected school year and semester
        eligible_students = []
        clearance_status_counts = {'cleared': 0, 'pending': 0, 'not_started': 0}

        for student in filtered_students:
            clearance = Clearance.objects.filter(
                student=student,
                school_year=selected_school_year,
                semester=selected_semester
            ).first()

            if clearance:
                if clearance.is_cleared:
                    clearance_status_counts['cleared'] += 1
                else:
                    clearance_status_counts['pending'] += 1
                eligible_students.append(student)
            else:
                clearance_status_counts['not_started'] += 1

        # Get recent batch approvals (last 10)
        # In a real implementation, this would come from a BatchApproval model
        # For now, we'll create a placeholder with sample data
        recent_batch_approvals = [
            {
                'id': 1,
                'date': timezone.now() - timezone.timedelta(days=1),
                'approved_by': request.user,
                'student_count': 15,
                'type': 'final'
            },
            {
                'id': 2,
                'date': timezone.now() - timezone.timedelta(days=3),
                'approved_by': request.user,
                'student_count': 8,
                'type': 'permit'
            }
        ]

        # Calculate percentages for the filter info
        total_students = len(filtered_students)
        if total_students > 0:
            cleared_percentage = int((clearance_status_counts['cleared'] / total_students) * 100)
            pending_percentage = int((clearance_status_counts['pending'] / total_students) * 100)
            not_started_percentage = int((clearance_status_counts['not_started'] / total_students) * 100)
        else:
            cleared_percentage = pending_percentage = not_started_percentage = 0

        # Add batch approval specific context
        context.update({
            'eligible_students': eligible_students,
            'recent_batch_approvals': recent_batch_approvals,
            'year_level': year_level,
            'cleared_count': clearance_status_counts['cleared'],
            'pending_count': clearance_status_counts['pending'],
            'not_started_count': clearance_status_counts['not_started'],
            'cleared_percentage': cleared_percentage,
            'pending_percentage': pending_percentage,
            'not_started_percentage': not_started_percentage,
            'total_students': total_students,
        })

    elif view == 'reports':
        # Calculate statistics for reports
        total_clearances = clearances.count() or 1  # Avoid division by zero
        cleared_count = clearances.filter(is_cleared=True).count()
        pending_count = clearances.filter(is_cleared=False).count()
        permits_issued = clearances.filter(is_cleared=True, program_chair_approved=True).count()

        cleared_percentage = int((cleared_count / total_clearances) * 100) if total_clearances > 0 else 0
        pending_percentage = int((pending_count / total_clearances) * 100) if total_clearances > 0 else 0
        permits_percentage = int((permits_issued / cleared_count) * 100) if cleared_count > 0 else 0

        # Add reports specific context
        context.update({
            'cleared_count': cleared_count,
            'pending_count': pending_count,
            'permits_issued': permits_issued,
            'cleared_percentage': cleared_percentage,
            'pending_percentage': pending_percentage,
            'permits_percentage': permits_percentage,
        })

    elif view == 'permit_printing':
        filter_type = request.GET.get('filter', '')
        search_query = request.GET.get('search', '')
        department_id = request.GET.get('department', '')
        course_id = request.GET.get('course', '')
        year_level = request.GET.get('year_level', '')

        # Get all cleared students
        cleared_student_ids = clearances.filter(is_cleared=True).values_list('student_id', flat=True)
        cleared_students_qs = students.filter(id__in=cleared_student_ids)

        # Get students with permits issued
        permit_issued_ids = clearances.filter(
            is_cleared=True,
            program_chair_approved=True
        ).values_list('student_id', flat=True)

        # Calculate permit statistics
        eligible_count = cleared_students_qs.count()
        issued_count = len(permit_issued_ids)
        pending_permit_count = eligible_count - issued_count

        issued_percentage = int((issued_count / eligible_count) * 100) if eligible_count > 0 else 0
        pending_permit_percentage = int((pending_permit_count / eligible_count) * 100) if eligible_count > 0 else 0

        # Filter students based on permit status
        filtered_students = students
        if filter_type == 'eligible':
            filtered_students = cleared_students_qs
        elif filter_type == 'issued':
            filtered_students = students.filter(id__in=permit_issued_ids)
        elif filter_type == 'pending':
            filtered_students = cleared_students_qs.exclude(id__in=permit_issued_ids)

        # Apply additional filters
        if department_id:
            filtered_students = filtered_students.filter(course__dean_id=department_id)
        if course_id:
            filtered_students = filtered_students.filter(course_id=course_id)
        if year_level:
            filtered_students = filtered_students.filter(year_level=year_level)

        # Apply search filter if provided
        if search_query:
            filtered_students = filtered_students.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(student_id__icontains=search_query)
            )

        # Get courses for the filter dropdown
        courses = Course.objects.filter(dean=program_chair.dean)

        # Get departments (deans) for the filter dropdown
        departments = Dean.objects.all()

        # Get recent permits (clearances with program_chair_approved=True)
        recent_permits = clearances.filter(
            is_cleared=True,
            program_chair_approved=True
        ).select_related(
            'student', 'student__user', 'student__course'
        ).order_by('-cleared_date')[:10]

        # Transform clearances into permit format for the template
        formatted_permits = []
        for clearance in recent_permits:
            formatted_permits.append({
                'created_at': clearance.cleared_date or clearance.created_at,
                'student': clearance.student,
                'issued_by': request.user,
                'school_year': clearance.school_year,
                'semester': clearance.semester
            })

        # Paginate the results
        paginator = Paginator(filtered_students, 10)
        page = request.GET.get('page')
        students_page = paginator.get_page(page)

        # Add permit printing specific context
        context.update({
            'students': students_page,
            'eligible_count': eligible_count,
            'issued_count': issued_count,
            'pending_permit_count': pending_permit_count,
            'issued_percentage': issued_percentage,
            'pending_permit_percentage': pending_permit_percentage,
            'courses': courses,
            'departments': departments,
            'recent_permits': formatted_permits,
        })

    else:
        # Default dashboard view
        # Get recent activities from ClearanceRequest model
        # Get students under this program chair's dean
        student_ids = students.values_list('id', flat=True)

        # Get recent clearance requests for these students
        recent_activities = ClearanceRequest.objects.filter(
            student_id__in=student_ids,
            status__in=['approved', 'denied']
        ).select_related(
            'student', 'student__user', 'office', 'reviewed_by'
        ).order_by('-reviewed_date')[:10]

        # Transform clearance requests into activity format for the template
        activities = []
        for clearance_req in recent_activities:
            status = clearance_req.status
            description = f"{clearance_req.office.name} clearance {status}"

            # Add remarks if available
            if clearance_req.remarks:
                description += f": {clearance_req.remarks}"

            activities.append({
                'date': clearance_req.reviewed_date or clearance_req.request_date,
                'student': clearance_req.student,
                'description': description,
                'status': status,
                'office': clearance_req.office.name,
                'reviewer': clearance_req.reviewed_by.get_full_name() if clearance_req.reviewed_by else 'System'
            })

        # Paginate students for the default view
        paginator = Paginator(students, 10)
        page = request.GET.get('page')
        students_page = paginator.get_page(page)

        # Add default dashboard specific context
        context.update({
            'students': students_page,
            'recent_activities': activities,
            'page_obj': students_page,
        })

    return render(request, 'core/program_chair_dashboard.html', context)

@login_required
@user_passes_test(is_program_chair)
def program_chair_profile(request):
    try:
        program_chair = request.user.programchair

        if request.method == 'POST' and request.POST.get('action') == 'change_password':
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            # Validate passwords
            if not request.user.check_password(current_password):
                messages.error(request, 'Current password is incorrect.')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match.')
            else:
                # Set new password with Django's built-in password hashing
                request.user.set_password(new_password)
                request.user.save()
                messages.success(request, 'Password changed successfully. Please log in with your new password.')
                return redirect('login')

        return render(request, 'core/program_chair_profile.html', {
            'program_chair': program_chair
        })
    except ProgramChair.DoesNotExist:
        messages.error(request, "Program chair profile not found.")
        return redirect('home')

@login_required
def delete_clearance(request, clearance_id):
    clearance = get_object_or_404(Clearance, id=clearance_id)
    if request.method == 'POST':
        clearance.delete()
        messages.success(request, 'Clearance deleted successfully.')
    return redirect('program_chair_dashboard')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_delete_clearance(request, clearance_id):
    """API endpoint for admin to delete a clearance with JSON response"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        clearance = get_object_or_404(Clearance, id=clearance_id)
        clearance.delete()
        return JsonResponse({'success': True, 'message': 'Clearance deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

class ManageStudentsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Student
    template_name = 'core/manage_students.html'
    context_object_name = 'students'
    paginate_by = 10

    def test_func(self):
        return hasattr(self.request.user, 'programchair')

    def get_queryset(self):
        program_chair = self.request.user.programchair
        queryset = Student.objects.filter(
            course__dean=program_chair.dean
        ).select_related('user', 'course').prefetch_related('clearances')

        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(student_id__icontains=search_query)
            )

        year_level = self.request.GET.get('year_level')
        if year_level:
            queryset = queryset.filter(year_level=year_level)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_year = timezone.now().year
        context.update({
            'program_chair': self.request.user.programchair,
            'current_school_year': f"{current_year}-{current_year + 1}",
            'current_semester': "1ST",
        })
        return context

    def handle_no_permission(self):
        messages.error(self.request, "Permission denied.")
        return redirect('home')

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        student_id = request.POST.get('student_id')

        try:
            student = Student.objects.get(id=student_id)
            if action == 'delete':
                student.user.delete()
                messages.success(request, 'Student deleted successfully.')
            elif action == 'approve':
                student.is_approved = True
                student.approval_date = timezone.now()
                student.approval_admin = request.user
                student.save()
                messages.success(request, 'Student approved successfully.')
        except Student.DoesNotExist:
            messages.error(request, 'Student not found.')

        return redirect('manage_students')

class StudentDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Student
    template_name = 'core/student_detail.html'
    context_object_name = 'student'

    def test_func(self):
        return hasattr(self.request.user, 'programchair')

    def get_queryset(self):
        return Student.objects.filter(
            course__dean=self.request.user.programchair.dean
        ).select_related('user', 'course', 'program_chair', 'program_chair__user')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object

        # Get the current clearance
        current_clearance = student.get_current_clearance()
        context['current_clearance'] = current_clearance

        # Get clearance requests for the current clearance
        if current_clearance:
            clearance_requests = ClearanceRequest.objects.filter(
                clearance=current_clearance
            ).select_related('office', 'reviewed_by')
            context['clearance_requests'] = clearance_requests

        return context

    def handle_no_permission(self):
        messages.error(self.request, "Permission denied.")
        return redirect('home')

# Staff Views
@login_required
def staff_dashboard(request):
    staff = request.user.staff

    # Check if the staff member is a dormitory owner and redirect to BH Owner Dashboard
    if staff.is_dormitory_owner:
        return redirect('bh_owner_dashboard')

    # Get current school year and semester
    current_year = timezone.now().year
    school_year = f"{current_year}-{current_year + 1}"
    # Use the utils.py version of get_current_semester that includes midterm/final
    from core.utils import get_current_semester as get_detailed_semester
    semester = get_detailed_semester()

    # Check if we're in a student management view
    view = request.GET.get('view', '')
    if 'student_management' in view:
        return handle_student_management(request, staff, school_year, semester)

    # Get recent requests for this office
    recent_requests = ClearanceRequest.objects.filter(
        office=staff.office
    ).select_related(
        'student__user',
        'student__course'
    ).order_by('-request_date')[:10]  # Get last 10 requests using request_date

    # Get pending requests count
    pending_requests_count = ClearanceRequest.objects.filter(
        office=staff.office,
        status='pending'
    ).count()

    # Get approved requests count for today
    today = timezone.now().date()
    approved_today_count = ClearanceRequest.objects.filter(
        office=staff.office,
        status='approved',
        reviewed_date__date=today
    ).count()

    # Get total processed count
    total_processed = ClearanceRequest.objects.filter(
        office=staff.office,
        status__in=['approved', 'denied']
    ).count()

    return render(request, 'core/staff_dashboard.html', {
        'recent_requests': recent_requests,
        'pending_requests_count': pending_requests_count,
        'approved_today_count': approved_today_count,
        'total_processed': total_processed,
        'school_year': school_year,
        'current_semester': semester,
        'office': staff.office,
    })

@login_required
def office_reports(request):
    staff = request.user.staff

    # Check if the staff member is a dormitory owner and redirect to BH Owner Dashboard
    if staff.is_dormitory_owner:
        return redirect('bh_owner_dashboard')

    # Get current school year and semester
    current_year = timezone.now().year
    school_year = f"{current_year}-{current_year + 1}"
    # Use the utils.py version of get_current_semester that includes midterm/final
    from core.utils import get_current_semester as get_detailed_semester
    semester = get_detailed_semester()

    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        school_year = request.POST.get('school_year')
        semester = request.POST.get('semester')

        if report_type == 'clearance_status':
            # Generate clearance status report
            clearance_requests = ClearanceRequest.objects.filter(
                office=staff.office,
                school_year=school_year,
                semester=semester
            ).select_related('student__user', 'reviewed_by')

            from core.office_pdf_utils import generate_office_clearance_report_pdf
            pdf = generate_office_clearance_report_pdf(clearance_requests, request, school_year, semester, staff.office)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="office_clearance_report_{school_year}_{semester}.pdf"'
            return response

        elif report_type == 'student_performance':
            # Generate student performance report
            clearance_requests = ClearanceRequest.objects.filter(
                office=staff.office,
                school_year=school_year,
                semester=semester
            ).select_related('student__user', 'student__course')

            from core.office_pdf_utils import generate_student_performance_report_pdf
            pdf = generate_student_performance_report_pdf(clearance_requests, request, school_year, semester, staff.office)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="student_performance_report_{school_year}_{semester}.pdf"'
            return response

        elif report_type == 'department_report':
            # Generate department report
            department_id = request.POST.get('department')

            # Base query for clearance requests from this office
            clearance_requests = ClearanceRequest.objects.filter(
                office=staff.office,
                school_year=school_year,
                semester=semester
            ).select_related('student__user', 'student__course', 'student__course__dean')

            # If a specific department is selected, filter by it
            if department_id:
                clearance_requests = clearance_requests.filter(student__course__dean_id=department_id)

            from core.office_pdf_utils import generate_department_report_pdf
            pdf = generate_department_report_pdf(clearance_requests, request, school_year, semester, staff.office, department_id)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="department_report_{school_year}_{semester}.pdf"'
            return response

        elif report_type == 'processing_time':
            # Generate processing time report
            clearance_requests = ClearanceRequest.objects.filter(
                office=staff.office,
                school_year=school_year,
                semester=semester,
                status__in=['approved', 'denied']
            ).select_related('student__user')

            from core.office_pdf_utils import generate_processing_time_report_pdf
            pdf = generate_processing_time_report_pdf(clearance_requests, request, school_year, semester, staff.office)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="processing_time_report_{school_year}_{semester}.pdf"'
            return response

        elif report_type == 'denial_reasons':
            # Generate denial reasons report
            clearance_requests = ClearanceRequest.objects.filter(
                office=staff.office,
                school_year=school_year,
                semester=semester,
                status='denied'
            ).select_related('student__user')

            from core.office_pdf_utils import generate_denial_reasons_report_pdf
            pdf = generate_denial_reasons_report_pdf(clearance_requests, request, school_year, semester, staff.office)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="denial_reasons_report_{school_year}_{semester}.pdf"'
            return response

        elif report_type == 'cleared_students':
            # Generate cleared students report
            clearance_requests = ClearanceRequest.objects.filter(
                office=staff.office,
                school_year=school_year,
                semester=semester,
                status='approved'
            ).select_related('student__user', 'student__course')

            from core.cleared_students_pdf import generate_cleared_students_report_pdf
            pdf = generate_cleared_students_report_pdf(clearance_requests, request, school_year, semester, staff.office)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="cleared_students_report_{school_year}_{semester}.pdf"'
            return response

    # Get all deans for department filter dropdown
    deans = Dean.objects.all().order_by('name')

    return render(request, 'core/office_reports.html', {
        'office': staff.office,
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
        'deans': deans,
    })

@login_required
def handle_student_management(request, staff, default_school_year, default_semester):
    view = request.GET.get('view', '')

    # Get selected school year and semester from request or use defaults
    selected_school_year = request.GET.get('school_year', default_school_year)
    selected_semester = request.GET.get('semester', default_semester)

    # Generate available years (current year and 4 years back)
    current_year = timezone.now().year
    available_years = [f"{year}-{year + 1}" for year in range(current_year - 4, current_year + 1)]

    # Common data for all student management views
    context = {
        'school_year': default_school_year,
        'current_semester': default_semester,
        'selected_school_year': selected_school_year,
        'selected_semester': selected_semester,
        'available_years': available_years,
        'office': staff.office,
        'semester_choices': SEMESTER_CHOICES,
    }

    # Get all courses for filtering
    courses = Course.objects.all().order_by('code')
    context['courses'] = courses

    # Handle specific student management views
    if view == 'student_management_all':
        # Get all students with pagination
        students_list = Student.objects.select_related('user', 'course').all()

        # Apply filters if provided
        search_query = request.GET.get('search', '')
        year_level = request.GET.get('year_level', '')
        course_id = request.GET.get('course', '')
        page_size = request.GET.get('page_size', 10)

        if search_query:
            students_list = students_list.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(student_id__icontains=search_query) |
                Q(contact_number__icontains=search_query)
            )

        if year_level:
            students_list = students_list.filter(year_level=year_level)

        if course_id:
            students_list = students_list.filter(course_id=course_id)

        # Paginate results
        paginator = Paginator(students_list, int(page_size))
        page = request.GET.get('page', 1)
        students = paginator.get_page(page)

        context['students'] = students
        context['page_sizes'] = [10, 25, 50, 100]

    elif view == 'student_management_search':
        # Handle search functionality
        search_term = request.GET.get('search_term', '')
        search_type = request.GET.get('search_type', 'all')
        year_level = request.GET.get('year_level', '')
        course_id = request.GET.get('course', '')
        clearance_status = request.GET.get('clearance_status', '')
        page_size = request.GET.get('page_size', 10)

        # Check if search was performed
        search_performed = bool(search_term or year_level or course_id or clearance_status)
        context['search_performed'] = search_performed

        if search_performed:
            # Start with all students
            search_results = Student.objects.select_related('user', 'course').all()

            # Apply search term filter based on search type
            if search_term:
                if search_type == 'name':
                    search_results = search_results.filter(
                        Q(user__first_name__icontains=search_term) |
                        Q(user__last_name__icontains=search_term)
                    )
                elif search_type == 'id':
                    search_results = search_results.filter(student_id__icontains=search_term)
                elif search_type == 'email':
                    search_results = search_results.filter(user__email__icontains=search_term)
                elif search_type == 'contact':
                    search_results = search_results.filter(contact_number__icontains=search_term)
                else:  # 'all'
                    search_results = search_results.filter(
                        Q(user__first_name__icontains=search_term) |
                        Q(user__last_name__icontains=search_term) |
                        Q(student_id__icontains=search_term) |
                        Q(user__email__icontains=search_term) |
                        Q(contact_number__icontains=search_term)
                    )

            # Apply additional filters
            if year_level:
                search_results = search_results.filter(year_level=year_level)

            if course_id:
                search_results = search_results.filter(course_id=course_id)

            if clearance_status:
                # Get current clearances for the current school year and semester
                if clearance_status == 'cleared':
                    search_results = search_results.filter(
                        clearances__school_year=selected_school_year,
                        clearances__semester=selected_semester,
                        clearances__is_cleared=True
                    ).distinct()
                else:  # 'pending'
                    search_results = search_results.filter(
                        Q(clearances__school_year=selected_school_year, clearances__semester=selected_semester, clearances__is_cleared=False) |
                        ~Q(clearances__school_year=selected_school_year, clearances__semester=selected_semester)
                    ).distinct()

            # Paginate results
            paginator = Paginator(search_results, int(page_size))
            page = request.GET.get('page', 1)
            paginated_results = paginator.get_page(page)

            context['search_results'] = paginated_results
            context['page_sizes'] = [10, 25, 50, 100]

    elif view == 'student_management_year':
        # Get student counts by year level
        year_counts = {}
        for year in range(1, 6):  # 1st to 5th year
            year_counts[year] = Student.objects.filter(year_level=year).count()

        context['year_counts'] = year_counts

        # If a year level is selected, get students for that year
        selected_year = request.GET.get('year_level')
        page_size = request.GET.get('page_size', 10)

        if selected_year:
            # Get all students for the selected year
            year_students_list = Student.objects.filter(year_level=selected_year).select_related('user', 'course')

            # Paginate results
            paginator = Paginator(year_students_list, int(page_size))
            page = request.GET.get('page', 1)
            year_students = paginator.get_page(page)

            context['year_students'] = year_students
            context['page_sizes'] = [10, 25, 50, 100]

    elif view == 'student_management_course':
        # Get student counts by course
        course_counts = {}
        for course in courses:
            course_counts[course.id] = Student.objects.filter(course=course).count()

        context['course_counts'] = course_counts

        # If a course is selected, get students for that course
        selected_course_id = request.GET.get('course_id')
        page_size = request.GET.get('page_size', 10)

        if selected_course_id:
            try:
                selected_course = Course.objects.get(id=selected_course_id)
                course_students_list = Student.objects.filter(course=selected_course).select_related('user')

                # Paginate results
                paginator = Paginator(course_students_list, int(page_size))
                page = request.GET.get('page', 1)
                course_students = paginator.get_page(page)

                context['selected_course'] = selected_course
                context['course_students'] = course_students
                context['page_sizes'] = [10, 25, 50, 100]
            except Course.DoesNotExist:
                pass

    elif view == 'student_management_clearance':
        # Get clearance statistics
        total_students = Student.objects.count()
        cleared_students = Student.objects.filter(
            clearances__school_year=selected_school_year,
            clearances__semester=selected_semester,
            clearances__is_cleared=True
        ).distinct().count()

        pending_students = total_students - cleared_students

        # Calculate percentages
        cleared_percentage = int((cleared_students / total_students) * 100) if total_students > 0 else 0
        pending_percentage = 100 - cleared_percentage

        # Get course-specific clearance stats
        course_stats = {}
        for course in courses:
            course_total = Student.objects.filter(course=course).count()
            course_cleared = Student.objects.filter(
                course=course,
                clearances__school_year=selected_school_year,
                clearances__semester=selected_semester,
                clearances__is_cleared=True
            ).distinct().count()

            course_pending = course_total - course_cleared
            course_cleared_percentage = int((course_cleared / course_total) * 100) if course_total > 0 else 0

            course_stats[course.id] = {
                'total': course_total,
                'cleared': course_cleared,
                'pending': course_pending,
                'cleared_percentage': course_cleared_percentage
            }

        context.update({
            'total_students': total_students,
            'cleared_count': cleared_students,
            'pending_count': pending_students,
            'cleared_percentage': cleared_percentage,
            'pending_percentage': pending_percentage,
            'course_stats': course_stats
        })

        # If a status is selected, get students with that status
        status = request.GET.get('status')
        page_size = request.GET.get('page_size', 10)

        if status:
            if status == 'cleared':
                status_students_list = Student.objects.filter(
                    clearances__school_year=selected_school_year,
                    clearances__semester=selected_semester,
                    clearances__is_cleared=True
                ).select_related('user', 'course').distinct()
            else:  # 'pending'
                # Get students who either have a non-cleared clearance or no clearance for the current period
                status_students_list = Student.objects.filter(
                    Q(clearances__school_year=selected_school_year, clearances__semester=selected_semester, clearances__is_cleared=False) |
                    ~Q(clearances__school_year=selected_school_year, clearances__semester=selected_semester)
                ).select_related('user', 'course').distinct()

            # Paginate results
            paginator = Paginator(status_students_list, int(page_size))
            page = request.GET.get('page', 1)
            status_students = paginator.get_page(page)

            context['status_students'] = status_students
            context['page_sizes'] = [10, 25, 50, 100]

    elif view == 'student_management_export':
        # Get available school years for export
        available_school_years = Clearance.objects.values_list('school_year', flat=True).distinct()
        context['available_school_years'] = available_school_years

        # Handle export form submissions
        if request.method == 'POST':
            action = request.POST.get('action')

            if action == 'export_all_students':
                # Logic for exporting all students
                export_format = request.POST.get('export_format', 'csv')
                # This would typically generate a file and return it as a response
                messages.success(request, f'All students exported successfully as {export_format.upper()}')

            elif action == 'export_filtered_students':
                # Logic for exporting filtered students
                year_level = request.POST.get('year_level')
                course_id = request.POST.get('course')
                clearance_status = request.POST.get('clearance_status')
                export_format = request.POST.get('export_format', 'csv')

                # This would typically generate a filtered file and return it
                messages.success(request, f'Filtered students exported successfully as {export_format.upper()}')

            elif action == 'export_clearance_report':
                # Logic for exporting clearance report
                report_school_year = request.POST.get('school_year', selected_school_year)
                report_semester = request.POST.get('semester', selected_semester)
                export_format = request.POST.get('export_format', 'pdf')

                # This would typically generate a report and return it
                messages.success(request, f'Clearance report for {report_school_year} {report_semester} exported successfully as {export_format.upper()}')

        # Mock recent exports for demonstration
        recent_exports = [
            {
                'filename': f'All_Students_{selected_school_year}_{selected_semester}.csv',
                'type': 'csv',
                'date_generated': timezone.now().strftime('%b %d, %Y'),
                'size': '42 KB',
                'download_url': '#'
            },
            {
                'filename': f'Clearance_Report_{selected_school_year}_{selected_semester}.pdf',
                'type': 'pdf',
                'date_generated': timezone.now().strftime('%b %d, %Y'),
                'size': '156 KB',
                'download_url': '#'
            }
        ]

        context['recent_exports'] = recent_exports

    elif view == 'student_management_batch_approval':
        # Handle AJAX requests for batch approval
        if request.GET.get('ajax') == '1':
            # Get filter parameters
            year_level = request.GET.get('year_level', '')
            course_id = request.GET.get('course', '')
            status = request.GET.get('status', 'pending')
            school_year = request.GET.get('school_year', selected_school_year)
            semester = request.GET.get('semester', selected_semester)

            # Get pagination parameters
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))

            # Start with all students
            students = Student.objects.select_related('user', 'course').all()

            # Apply filters
            if year_level:
                students = students.filter(year_level=year_level)

            if course_id:
                students = students.filter(course_id=course_id)

            # Filter by clearance status
            if status == 'pending':
                students = students.filter(
                    Q(clearances__school_year=school_year, clearances__semester=semester, clearances__is_cleared=False) |
                    ~Q(clearances__school_year=school_year, clearances__semester=semester)
                ).distinct()
            elif status == 'cleared':
                students = students.filter(
                    clearances__school_year=school_year,
                    clearances__semester=semester,
                    clearances__is_cleared=True
                ).distinct()

            # Get total count for pagination
            total_count = students.count()
            total_pages = (total_count + page_size - 1) // page_size  # Ceiling division

            # Apply pagination
            start_index = (page - 1) * page_size
            end_index = start_index + page_size
            paginated_students = students[start_index:end_index]

            # Prepare data for JSON response
            students_data = []
            for student in paginated_students:
                clearance = Clearance.objects.filter(
                    student=student,
                    school_year=school_year,
                    semester=semester
                ).first()

                students_data.append({
                    'id': student.id,
                    'full_name': student.user.get_full_name(),
                    'email': student.user.email,
                    'student_id': student.student_id,
                    'course_code': student.course.code if student.course else '',
                    'year_level': student.year_level,
                    'is_cleared': clearance.is_cleared if clearance else False
                })

            # Return paginated data with pagination metadata
            return JsonResponse({
                'students': students_data,
                'pagination': {
                    'total_count': total_count,
                    'total_pages': total_pages,
                    'current_page': page,
                    'page_size': page_size,
                    'has_next': page < total_pages,
                    'has_previous': page > 1
                }
            })

        # Handle AJAX requests for preview
        elif request.GET.get('ajax') == '2':
            selected_students = request.GET.get('selected_students', '')
            if selected_students:
                student_ids = [int(id) for id in selected_students.split(',')]
                students = Student.objects.filter(id__in=student_ids).select_related('user', 'course')

                # Prepare data for JSON response
                students_data = [{
                    'id': student.id,
                    'full_name': student.user.get_full_name(),
                    'student_id': student.student_id,
                    'course_code': student.course.code if student.course else '',
                    'year_level': student.year_level
                } for student in students]

                return JsonResponse({'students': students_data})
            return JsonResponse({'students': []})

        # Handle POST request for batch approval
        if request.method == 'POST' and request.POST.get('action') == 'batch_approve':
            selected_students = request.POST.getlist('selected_students[]')
            approval_type = request.POST.get('approval_type')
            semester = request.POST.get('semester')
            comments = request.POST.get('comments')

            if selected_students:
                # Process batch approval
                approved_count = 0
                for student_id in selected_students:
                    try:
                        student = Student.objects.get(id=student_id)
                        clearance, created = Clearance.objects.get_or_create(
                            student=student,
                            school_year=selected_school_year,
                            semester=semester,
                            defaults={'is_cleared': False}
                        )

                        if approval_type == 'full':
                            # Full clearance - mark all requests as approved
                            clearance_requests = ClearanceRequest.objects.filter(
                                clearance=clearance,
                                status='pending'
                            )

                            for req in clearance_requests:
                                req.status = 'approved'
                                req.reviewed_by = staff
                                req.reviewed_date = timezone.now()
                                req.notes = comments if comments else None
                                req.save()

                            clearance.is_cleared = True
                            clearance.cleared_date = timezone.now()
                            clearance.save()
                            approved_count += 1

                        elif approval_type == 'partial':
                            # Partial clearance - mark only this office's requests as approved
                            clearance_request = ClearanceRequest.objects.filter(
                                clearance=clearance,
                                office=staff.office,
                                status='pending'
                            ).first()

                            if clearance_request:
                                clearance_request.status = 'approved'
                                clearance_request.reviewed_by = staff
                                clearance_request.reviewed_date = timezone.now()
                                clearance_request.notes = comments if comments else None
                                clearance_request.save()

                                # Check if all requests are now approved
                                pending_requests = ClearanceRequest.objects.filter(
                                    clearance=clearance,
                                    status='pending'
                                ).exists()

                                if not pending_requests:
                                    clearance.is_cleared = True
                                    clearance.cleared_date = timezone.now()
                                    clearance.save()

                                approved_count += 1

                    except Student.DoesNotExist:
                        continue

                # Create a record of this batch approval
                batch_approval = {
                    'id': str(timezone.now().timestamp()),
                    'date': timezone.now(),
                    'approved_by': staff.user.get_full_name(),
                    'student_count': approved_count,
                    'type': 'Full Clearance' if approval_type == 'full' else 'Partial Clearance',
                    'semester': dict(SEMESTER_CHOICES).get(semester, semester),
                    'school_year': selected_school_year,
                    'comments': comments
                }

                # Store batch approval in session
                batch_approvals = request.session.get('batch_approvals', [])
                batch_approvals.insert(0, batch_approval)
                request.session['batch_approvals'] = batch_approvals[:10]  # Keep only the 10 most recent

                messages.success(request, f'Successfully approved {approved_count} students.')
                return redirect('staff_dashboard')

        # Get recent batch approvals from session
        recent_batch_approvals = request.session.get('batch_approvals', [])
        context['recent_batch_approvals'] = recent_batch_approvals

    elif view == 'student_management_clearance_review':
        # Get pending clearance requests for this staff member's office
        pending_requests = ClearanceRequest.objects.filter(
            office=staff.office,
            status='pending'
        ).select_related('student__user', 'student__course').order_by('-request_date')

        # Get recent activities
        recent_activities = ClearanceRequest.objects.filter(
            office=staff.office,
            status__in=['approved', 'denied']
        ).select_related('student__user', 'reviewed_by').order_by('-reviewed_date')[:10]

        context.update({
            'pending_requests': pending_requests,
            'recent_activities': recent_activities
        })

    elif view == 'student_management_analytics':
        # Get clearance statistics
        total_students = Student.objects.count()
        cleared_students = Student.objects.filter(
            clearances__school_year=selected_school_year,
            clearances__semester=selected_semester,
            clearances__is_cleared=True
        ).distinct().count()

        pending_students = total_students - cleared_students

        # Calculate percentages
        cleared_percentage = int((cleared_students / total_students) * 100) if total_students > 0 else 0
        pending_percentage = 100 - cleared_percentage

        # Get course-specific clearance stats
        course_stats = {}
        for course in courses:
            course_total = Student.objects.filter(course=course).count()
            course_cleared = Student.objects.filter(
                course=course,
                clearances__school_year=selected_school_year,
                clearances__semester=selected_semester,
                clearances__is_cleared=True
            ).distinct().count()

            course_pending = course_total - course_cleared
            course_cleared_percentage = int((course_cleared / course_total) * 100) if course_total > 0 else 0

            course_stats[course.id] = {
                'total': course_total,
                'cleared': course_cleared,
                'pending': course_pending,
                'cleared_percentage': course_cleared_percentage
            }

        # Get year level stats
        year_stats = {}
        year_levels = range(1, 6)  # 1st to 5th year
        for year in year_levels:
            year_total = Student.objects.filter(year_level=year).count()
            year_cleared = Student.objects.filter(
                year_level=year,
                clearances__school_year=selected_school_year,
                clearances__semester=selected_semester,
                clearances__is_cleared=True
            ).distinct().count()

            year_pending = year_total - year_cleared
            year_cleared_percentage = int((year_cleared / year_total) * 100) if year_total > 0 else 0

            year_stats[year] = {
                'total': year_total,
                'cleared': year_cleared,
                'pending': year_pending,
                'cleared_percentage': year_cleared_percentage
            }

        # Calculate average processing time
        avg_processing_days = 3  # Mock data - would be calculated from actual clearance requests

        context.update({
            'total_students': total_students,
            'cleared_count': cleared_students,
            'pending_count': pending_students,
            'cleared_percentage': cleared_percentage,
            'pending_percentage': pending_percentage,
            'course_stats': course_stats,
            'year_stats': year_stats,
            'year_levels': year_levels,
            'avg_processing_days': avg_processing_days,
            'current_semester_display': dict(SEMESTER_CHOICES).get(selected_semester, selected_semester)
        })

    elif view == 'student_management_view':
        # View a specific student's details
        student_id = request.GET.get('student_id')
        if student_id:
            try:
                student = Student.objects.select_related('user', 'course', 'program_chair', 'dormitory_owner').get(id=student_id)

                # Get current clearance
                current_clearance = Clearance.objects.filter(
                    student=student,
                    school_year=selected_school_year,
                    semester=selected_semester
                ).first()

                # Get clearance requests if there's a current clearance
                clearance_requests = []
                clearance_percentage = 0
                if current_clearance:
                    clearance_requests = ClearanceRequest.objects.filter(
                        clearance=current_clearance
                    ).select_related('office', 'reviewed_by')

                    # Calculate clearance percentage
                    total_requests = clearance_requests.count()
                    approved_requests = clearance_requests.filter(status='approved').count()
                    clearance_percentage = int((approved_requests / total_requests) * 100) if total_requests > 0 else 0

                # Get clearance history
                clearance_history = Clearance.objects.filter(student=student).exclude(
                    school_year=selected_school_year, semester=selected_semester
                ).order_by('-school_year', '-semester')

                context.update({
                    'student': student,
                    'current_clearance': current_clearance,
                    'clearance_requests': clearance_requests,
                    'clearance_percentage': clearance_percentage,
                    'clearance_history': clearance_history
                })
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')

    return render(request, 'core/staff_student_management.html', context)

# Dormitory Owner Views
@login_required
def bh_owner_dashboard(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access the Dormitory Owner Dashboard.")
        return redirect('home')

    staff = request.user.staff

    # Get current school year and semester
    current_year = timezone.now().year
    school_year = f"{current_year}-{current_year + 1}"
    semester = get_current_semester()

    # Get clearance requests for students assigned to this dormitory owner
    clearance_requests = ClearanceRequest.objects.filter(
        student__dormitory_owner=staff,
        office__name="DORMITORY"
    ).select_related(
        'student__user',
        'student__course'
    ).order_by('-request_date')

    # Get counts for dashboard stats
    total_boarders = Student.objects.filter(dormitory_owner=staff).count()
    pending_requests = clearance_requests.filter(status='pending').count()
    approved_requests = clearance_requests.filter(status='approved').count()
    denied_requests = clearance_requests.filter(status='denied').count()

    return render(request, 'core/bh_owner_dashboard.html', {
        'clearance_requests': clearance_requests,
        'school_year': school_year,
        'current_semester': semester,
        'dormitory_owner': staff,
        'total_boarders': total_boarders,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'denied_requests': denied_requests,
    })

@login_required
def bh_owner_boarders(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    # Get all students assigned to this dormitory owner
    students = Student.objects.filter(dormitory_owner=staff).select_related('user', 'course')

    # Apply search filter if provided
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(student_id__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(students, 10)  # Show 10 students per page
    page = request.GET.get('page')
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        students_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        students_page = paginator.page(paginator.num_pages)

    return render(request, 'core/bh_owner_boarders.html', {
        'students': students_page,
        'dormitory_owner': staff,
        'search_query': search_query,
    })

@login_required
def update_boarder_date(request, student_id):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    staff = request.user.staff

    try:
        # Get the student and verify they belong to this dormitory owner
        student = Student.objects.get(id=student_id, dormitory_owner=staff)

        if request.method == 'POST':
            # Handle boarder_since date
            boarder_since = request.POST.get('boarder_since')
            if boarder_since:
                # Convert the datetime-local input to a Python datetime
                try:
                    from datetime import datetime
                    student.boarder_since = datetime.strptime(boarder_since, '%Y-%m-%dT%H:%M')
                except ValueError:
                    # If there's an error parsing the date, use current time
                    student.boarder_since = timezone.now()
            else:
                # If no date provided, use current time
                student.boarder_since = timezone.now()

            student.save()
            messages.success(request, f"Boarding start date for {student.get_full_name()} has been updated successfully.")

        return redirect('bh_owner_boarders')

    except Student.DoesNotExist:
        messages.error(request, "Student not found or you don't have permission to edit this student.")
        return redirect('bh_owner_boarders')

@login_required
def toggle_boarder_status(request, student_id):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    staff = request.user.staff

    try:
        # Get the student and verify they belong to this dormitory owner
        student = Student.objects.get(id=student_id, dormitory_owner=staff)

        if request.method == 'POST':
            # Toggle the boarder status
            student.is_boarder = not student.is_boarder

            # If student is becoming a boarder and doesn't have a boarder_since date, set it
            if student.is_boarder and not student.boarder_since:
                student.boarder_since = timezone.now()

            student.save()

            status_message = "activated" if student.is_boarder else "deactivated"
            messages.success(request, f"Boarder status for {student.get_full_name()} has been {status_message} successfully.")

        return redirect('bh_owner_boarders')

    except Student.DoesNotExist:
        messages.error(request, "Student not found or you don't have permission to edit this student.")
        return redirect('bh_owner_boarders')

@login_required
def bh_owner_add_student(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        try:
            # Find the student by ID
            student = Student.objects.get(student_id=student_id)

            # Check if student is already assigned to a dormitory owner
            if student.dormitory_owner:
                if student.dormitory_owner == staff:
                    messages.warning(request, f"Student {student.get_full_name()} is already assigned to your boarding house.")
                else:
                    messages.error(request, f"Student {student.get_full_name()} is already assigned to {student.dormitory_owner.get_full_name()}'s boarding house. Students cannot be in multiple boarding houses simultaneously.")
            else:
                # Assign student to this dormitory owner
                student.dormitory_owner = staff
                student.is_boarder = True

                # Handle boarder_since date
                boarder_since = request.POST.get('boarder_since')
                if boarder_since:
                    # Convert the datetime-local input to a Python datetime
                    try:
                        from datetime import datetime
                        student.boarder_since = datetime.strptime(boarder_since, '%Y-%m-%dT%H:%M')
                    except ValueError:
                        # If there's an error parsing the date, use current time
                        student.boarder_since = timezone.now()
                else:
                    # If no date provided, use current time
                    student.boarder_since = timezone.now()

                student.save()
                messages.success(request, f"Student {student.get_full_name()} has been added to your boarders.")
                return redirect('bh_owner_boarders')
        except Student.DoesNotExist:
            messages.error(request, f"No student found with ID {student_id}.")

    return render(request, 'core/bh_owner_add_student.html', {
        'dormitory_owner': staff,
    })

@login_required
def bh_owner_pending_requests(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    # Get pending clearance requests for students assigned to this dormitory owner
    clearance_requests = ClearanceRequest.objects.filter(
        student__dormitory_owner=staff,
        office__name="DORMITORY",
        status='pending'
    ).select_related(
        'student__user',
        'student__course'
    ).order_by('-request_date')

    # Apply search filter if provided
    search_query = request.GET.get('search', '')
    if search_query:
        clearance_requests = clearance_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(clearance_requests, 10)  # Show 10 requests per page
    page = request.GET.get('page')
    try:
        requests_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        requests_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        requests_page = paginator.page(paginator.num_pages)

    return render(request, 'core/bh_owner_pending_requests.html', {
        'clearance_requests': requests_page,
        'dormitory_owner': staff,
        'search_query': search_query,
    })

@login_required
def bh_owner_approved_requests(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    # Get approved clearance requests for students assigned to this dormitory owner
    clearance_requests = ClearanceRequest.objects.filter(
        student__dormitory_owner=staff,
        office__name="DORMITORY",
        status='approved'
    ).select_related(
        'student__user',
        'student__course',
        'reviewed_by'
    ).order_by('-reviewed_date')

    # Apply search filter if provided
    search_query = request.GET.get('search', '')
    if search_query:
        clearance_requests = clearance_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(clearance_requests, 10)  # Show 10 requests per page
    page = request.GET.get('page')
    try:
        requests_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        requests_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        requests_page = paginator.page(paginator.num_pages)

    return render(request, 'core/bh_owner_approved_requests.html', {
        'clearance_requests': requests_page,
        'dormitory_owner': staff,
        'search_query': search_query,
    })

@login_required
def bh_owner_denied_requests(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    # Get denied clearance requests for students assigned to this dormitory owner
    clearance_requests = ClearanceRequest.objects.filter(
        student__dormitory_owner=staff,
        office__name="DORMITORY",
        status='denied'
    ).select_related(
        'student__user',
        'student__course',
        'reviewed_by'
    ).order_by('-reviewed_date')

    # Apply search filter if provided
    search_query = request.GET.get('search', '')
    if search_query:
        clearance_requests = clearance_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(clearance_requests, 10)  # Show 10 requests per page
    page = request.GET.get('page')
    try:
        requests_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        requests_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        requests_page = paginator.page(paginator.num_pages)

    return render(request, 'core/bh_owner_denied_requests.html', {
        'clearance_requests': requests_page,
        'dormitory_owner': staff,
        'search_query': search_query,
    })

@login_required
def bh_owner_clearance_history(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    # Get all clearance requests for students assigned to this dormitory owner
    clearance_requests = ClearanceRequest.objects.filter(
        student__dormitory_owner=staff,
        office__name="DORMITORY"
    ).select_related(
        'student__user',
        'student__course',
        'reviewed_by'
    ).order_by('-request_date')

    # Apply filters if provided
    status = request.GET.get('status', '')
    if status:
        clearance_requests = clearance_requests.filter(status=status)

    school_year = request.GET.get('school_year', '')
    if school_year:
        clearance_requests = clearance_requests.filter(school_year=school_year)

    semester = request.GET.get('semester', '')
    if semester:
        clearance_requests = clearance_requests.filter(semester=semester)

    search_query = request.GET.get('search', '')
    if search_query:
        clearance_requests = clearance_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(clearance_requests, 10)  # Show 10 requests per page
    page = request.GET.get('page')
    try:
        requests_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        requests_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        requests_page = paginator.page(paginator.num_pages)

    return render(request, 'core/bh_owner_clearance_history.html', {
        'clearance_requests': requests_page,
        'dormitory_owner': staff,
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
        'current_filters': {
            'status': status,
            'school_year': school_year,
            'semester': semester,
            'search': search_query
        },
    })

@login_required
def bh_owner_generate_reports(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        school_year = request.POST.get('school_year')
        semester = request.POST.get('semester')

        # Get students based on filters
        students = Student.objects.filter(dormitory_owner=staff)

        if report_type == 'boarders':
            # Generate boarders report
            from core.office_pdf_utils import generate_boarders_pdf
            pdf = generate_boarders_pdf(students, request, school_year, semester)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="boarders_report_{school_year}_{semester}.pdf"'
            return response

        elif report_type == 'clearance':
            # Generate clearance report
            clearance_requests = ClearanceRequest.objects.filter(
                student__dormitory_owner=staff,
                office__name="DORMITORY",
                school_year=school_year,
                semester=semester
            )
            from core.office_pdf_utils import generate_clearance_report_pdf
            pdf = generate_clearance_report_pdf(clearance_requests, request, school_year, semester)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="clearance_report_{school_year}_{semester}.pdf"'
            return response

    return render(request, 'core/bh_owner_generate_reports.html', {
        'dormitory_owner': staff,
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
    })

@login_required
def bh_owner_export_data(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff

    if request.method == 'POST':
        export_type = request.POST.get('export_type')
        school_year = request.POST.get('school_year')
        semester = request.POST.get('semester')

        if export_type == 'boarders':
            # Export boarders to Excel
            students = Student.objects.filter(dormitory_owner=staff)

            # Create workbook and add worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Boarders"

            # Add headers
            headers = ['Student ID', 'Name', 'Course', 'Year Level', 'Contact Number']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0EBF5", end_color="E0EBF5", fill_type="solid")

            # Add data
            for row_num, student in enumerate(students, 2):
                ws.cell(row=row_num, column=1).value = student.student_id
                ws.cell(row=row_num, column=2).value = student.get_full_name()
                ws.cell(row=row_num, column=3).value = student.course.name if student.course else ''
                ws.cell(row=row_num, column=4).value = student.year_level
                ws.cell(row=row_num, column=5).value = student.contact_number

            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=boarders.xlsx'
            wb.save(response)
            return response

        elif export_type == 'clearance':
            # Export clearance data to Excel
            clearance_requests = ClearanceRequest.objects.filter(
                student__dormitory_owner=staff,
                office__name="DORMITORY",
                school_year=school_year,
                semester=semester
            )

            # Create workbook and add worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Clearance Requests"

            # Add headers
            headers = ['Student ID', 'Name', 'Request Date', 'Status', 'Reviewed Date', 'Reviewed By', 'Notes']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0EBF5", end_color="E0EBF5", fill_type="solid")

            # Add data
            for row_num, req in enumerate(clearance_requests, 2):
                ws.cell(row=row_num, column=1).value = req.student.student_id
                ws.cell(row=row_num, column=2).value = req.student.get_full_name()
                ws.cell(row=row_num, column=3).value = req.request_date.strftime('%Y-%m-%d %H:%M') if req.request_date else ''
                ws.cell(row=row_num, column=4).value = req.status.upper()
                ws.cell(row=row_num, column=5).value = req.reviewed_date.strftime('%Y-%m-%d %H:%M') if req.reviewed_date else ''
                ws.cell(row=row_num, column=6).value = req.reviewed_by.get_full_name() if req.reviewed_by else ''
                ws.cell(row=row_num, column=7).value = req.notes or ''

            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=clearance_requests_{school_year}_{semester}.xlsx'
            wb.save(response)
            return response

    return render(request, 'core/bh_owner_export_data.html', {
        'dormitory_owner': staff,
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
    })

@login_required
def update_clearance_request(request, request_id):
    """Handle updating clearance request status (approve/deny) for dormitory owners"""
    # Log the request for debugging
    logger.info(f"update_clearance_request called for request_id={request_id} by user_id={request.user.id}")
    logger.info(f"POST data: {request.POST}")

    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('home')

    staff = request.user.staff
    clearance_request = get_object_or_404(ClearanceRequest, id=request_id)

    # Log the clearance request details
    logger.info(f"Found clearance request: {clearance_request.id} for student {clearance_request.student.id}")

    # Check if this request belongs to a student assigned to this dormitory owner
    if clearance_request.student.dormitory_owner != staff:
        messages.error(request, "You don't have permission to update this clearance request.")
        return redirect('bh_owner_dashboard')

    # Check if the request is for the DORMITORY office
    if clearance_request.office.name != "DORMITORY":
        messages.error(request, "You can only update dormitory clearance requests.")
        return redirect('bh_owner_dashboard')

    # Check if request is still pending
    if clearance_request.status != 'pending':
        messages.warning(request, f'This request has already been {clearance_request.status}.')
        return redirect('bh_owner_pending_requests')

    if request.method == 'POST':
        status = request.POST.get('status')
        logger.info(f"Processing POST request with status={status}")

        if status not in ['approved', 'denied']:
            messages.error(request, "Invalid status.")
            return redirect('bh_owner_pending_requests')

        try:
            # Update the clearance request
            clearance_request.status = status
            clearance_request.reviewed_by = staff
            clearance_request.reviewed_date = timezone.now()

            # If denied, add the reason
            if status == 'denied':
                remarks = request.POST.get('remarks')
                if not remarks:
                    messages.error(request, "You must provide a reason for denial.")
                    return redirect('bh_owner_pending_requests')
                clearance_request.notes = remarks
            else:
                # Clear notes on approval
                clearance_request.notes = None

            clearance_request.save()
            logger.info(f"Clearance request {clearance_request.id} updated to status={status}")

            # Update the parent clearance
            clearance = Clearance.objects.get(
                student=clearance_request.student,
                school_year=clearance_request.school_year,
                semester=clearance_request.semester
            )
            clearance.check_clearance()

            # Show success message
            student_name = clearance_request.student.get_full_name()
            if status == 'approved':
                messages.success(request, f"Successfully approved clearance for {student_name}")
            else:
                messages.success(request, f"Clearance for {student_name} has been denied")

            # Determine where to redirect based on the referrer
            referer = request.META.get('HTTP_REFERER', '')
            if 'dashboard' in referer:
                logger.info(f"Redirecting to dashboard")
                return redirect('bh_owner_dashboard')
            else:
                logger.info(f"Redirecting to pending requests")
                return redirect('bh_owner_pending_requests')

        except Exception as e:
            logger.error(f"Error updating clearance request: {str(e)}", exc_info=True)
            messages.error(request, f"Error: {str(e)}")
            return redirect('bh_owner_pending_requests')

    # If not POST, redirect to pending requests
    return redirect('bh_owner_pending_requests')

@login_required
def bh_owner_profile(request):
    # Check if the user is a dormitory owner
    if not hasattr(request.user, 'staff') or not request.user.staff.is_dormitory_owner:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('home')

    staff = request.user.staff
    user = request.user

    if request.method == 'POST':
        # Update profile information
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.email = request.POST.get('email')
        user.save()

        # Update boarding house address
        staff.boarding_house_address = request.POST.get('boarding_house_address', '')
        staff.save()

        # Update password if provided
        new_password = request.POST.get('new_password')
        if new_password:
            user.set_password(new_password)
            user.save()
            messages.success(request, "Your profile has been updated. Please log in with your new password.")
            return redirect('login')

        messages.success(request, "Your profile has been updated successfully.")
        return redirect('bh_owner_profile')

    return render(request, 'core/bh_owner_profile.html', {
        'dormitory_owner': staff,
        'user': user,
    })

@login_required
def staff_profile(request):
    try:
        staff = request.user.staff

        if request.method == 'POST':
            if 'update_role' in request.POST:
                staff.role = request.POST.get('role')
                staff.save()
                messages.success(request, 'Role updated successfully')
                return redirect('staff_profile')

            elif request.POST.get('action') == 'change_password':
                current_password = request.POST.get('current_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')

                # Validate passwords
                if not request.user.check_password(current_password):
                    messages.error(request, 'Current password is incorrect.')
                elif new_password != confirm_password:
                    messages.error(request, 'New passwords do not match.')
                else:
                    # Set new password with Django's built-in password hashing
                    request.user.set_password(new_password)
                    request.user.save()
                    messages.success(request, 'Password changed successfully. Please log in with your new password.')
                    return redirect('login')

        return render(request, 'core/staff_profile.html', {
            'staff': staff,
            'office': staff.office,
            'user': request.user,
            'is_dormitory_owner': staff.is_dormitory_owner,
            'assigned_students': staff.students_dorm.all() if staff.is_dormitory_owner else None,
        })
    except Staff.DoesNotExist:
        messages.error(request, "Staff profile not found.")
        return redirect('home')

@login_required
def staff_pending_requests(request):
    staff = request.user.staff
    pending_requests = ClearanceRequest.objects.filter(
        office=staff.office, status='pending'
    )

    search_query = request.GET.get('search', '')
    if search_query:
        pending_requests = pending_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    school_year = request.GET.get('school_year', '')
    if school_year:
        pending_requests = pending_requests.filter(school_year=school_year)

    semester = request.GET.get('semester', '')
    if semester:
        pending_requests = pending_requests.filter(semester=semester)

    sort_by = request.GET.get('sort', 'request_date')
    if sort_by == 'student_name':
        pending_requests = pending_requests.order_by('student__user__last_name', 'student__user__first_name')
    elif sort_by == 'course':
        pending_requests = pending_requests.order_by('student__course__code')
    else:
        pending_requests = pending_requests.order_by('-request_date')

    return render(request, 'core/staff_pending_requests.html', {
        'pending_requests': pending_requests,
        'school_years': ClearanceRequest.objects.values_list('school_year', flat=True).distinct(),
        'SEMESTER_CHOICES': SEMESTER_CHOICES,
        'current_filters': {'search': search_query, 'school_year': school_year, 'semester': semester, 'sort': sort_by},
        'office': staff.office,
        'school_year': get_current_school_year(),
        'current_semester': get_current_semester(),
    })

@login_required
@require_POST
def approve_clearance_request(request, request_id):
    try:
        # Get the staff member and clearance request
        staff = request.user.staff
        clearance_request = get_object_or_404(ClearanceRequest, id=request_id)

        # Log the request for debugging
        logger.info(f"Approve request received for request_id={request_id} by staff={staff.id}")

        # Determine if user is a dormitory owner or regular staff
        is_dormitory_owner = staff.is_dormitory_owner
        redirect_url = 'bh_owner_pending_requests' if is_dormitory_owner else 'staff_pending_requests'

        # Check permissions based on user type
        if is_dormitory_owner:
            # For dormitory owners, check if this is their student and a dormitory request
            if clearance_request.student.dormitory_owner != staff:
                messages.error(request, "You don't have permission to update this clearance request.")
                return redirect(redirect_url)

            if clearance_request.office.name != "DORMITORY":
                messages.error(request, "You can only update dormitory clearance requests.")
                return redirect(redirect_url)
        else:
            # For regular staff, check if they belong to the right office
            if staff.office != clearance_request.office:
                messages.error(request, f"No permission for {clearance_request.office.name}")
                return redirect(redirect_url)

        # Check if request is still pending
        if clearance_request.status != 'pending':
            messages.warning(request, f'Request already processed (current status: {clearance_request.status})')
            return redirect(redirect_url)

        # Process the approval
        try:
            # Update the clearance request status
            clearance_request.status = "approved"
            clearance_request.reviewed_by = staff
            clearance_request.reviewed_date = timezone.now()
            clearance_request.notes = None  # Clear notes on approval
            clearance_request.save()

            # Update the parent clearance
            clearance = Clearance.objects.get(
                student=clearance_request.student,
                school_year=clearance_request.school_year,
                semester=clearance_request.semester
            )
            clearance.check_clearance()

            logger.info(f"Successfully approved request_id={request_id}")

            # Show success message and redirect
            student_name = clearance_request.student.get_full_name()
            messages.success(request, f'Successfully approved clearance for {student_name}')

            # Determine where to redirect based on the referrer
            referer = request.META.get('HTTP_REFERER', '')
            if is_dormitory_owner and 'dashboard' in referer:
                return redirect('bh_owner_dashboard')
            return redirect(redirect_url)

        except Exception as inner_e:
            logger.error(f"Error in approval process: {str(inner_e)}")
            messages.error(request, f'Error processing approval: {str(inner_e)}')
            return redirect(redirect_url)

    except Staff.DoesNotExist:
        logger.error(f"Staff access required for user_id={request.user.id}")
        messages.error(request, 'Staff access required')
        return redirect('login')

    except Exception as e:
        logger.error(f"Unexpected error in approve_clearance_request: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
        return redirect('staff_pending_requests')

@login_required
@require_POST
def deny_clearance_request(request, request_id):
    try:
        # Get the staff member and clearance request
        staff = request.user.staff
        clearance_request = get_object_or_404(ClearanceRequest, id=request_id)

        # Log the request for debugging
        logger.info(f"Deny request received for request_id={request_id} by staff={staff.id}")

        # Determine if user is a dormitory owner or regular staff
        is_dormitory_owner = staff.is_dormitory_owner
        redirect_url = 'bh_owner_pending_requests' if is_dormitory_owner else 'staff_pending_requests'

        # Check permissions based on user type
        if is_dormitory_owner:
            # For dormitory owners, check if this is their student and a dormitory request
            if clearance_request.student.dormitory_owner != staff:
                messages.error(request, "You don't have permission to update this clearance request.")
                return redirect(redirect_url)

            if clearance_request.office.name != "DORMITORY":
                messages.error(request, "You can only update dormitory clearance requests.")
                return redirect(redirect_url)
        else:
            # For regular staff, check if they belong to the right office
            if staff.office != clearance_request.office:
                messages.error(request, f"No permission for {clearance_request.office.name}")
                return redirect(redirect_url)

        # Check if request is still pending
        if clearance_request.status != 'pending':
            messages.warning(request, f'Request already processed (current status: {clearance_request.status})')
            return redirect(redirect_url)

        # Get the reason from form data - check combined_remarks first, then fall back to remarks
        reason = request.POST.get('combined_remarks', '') or request.POST.get('remarks', '')

        if not reason:
            messages.error(request, 'Reason required for denial')
            return redirect(redirect_url)

        try:
            # Update the clearance request status
            clearance_request.status = "denied"
            clearance_request.reviewed_by = staff
            clearance_request.reviewed_date = timezone.now()
            clearance_request.notes = reason
            clearance_request.save()

            # Update the parent clearance
            clearance = Clearance.objects.get(
                student=clearance_request.student,
                school_year=clearance_request.school_year,
                semester=clearance_request.semester
            )
            clearance.check_clearance()

            logger.info(f"Successfully denied request_id={request_id}")

            # Show success message and redirect
            student_name = clearance_request.student.get_full_name()
            messages.success(request, f'Successfully denied clearance for {student_name}')

            # Determine where to redirect based on the referrer
            referer = request.META.get('HTTP_REFERER', '')
            if is_dormitory_owner and 'dashboard' in referer:
                return redirect('bh_owner_dashboard')
            return redirect(redirect_url)

        except Exception as inner_e:
            logger.error(f"Error in denial process: {str(inner_e)}")
            messages.error(request, f'Error processing denial: {str(inner_e)}')
            return redirect(redirect_url)

    except Staff.DoesNotExist:
        logger.error(f"Staff access required for user_id={request.user.id}")
        messages.error(request, 'Staff access required')
        return redirect('login')

    except Exception as e:
        logger.error(f"Unexpected error in deny_clearance_request: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
        return redirect('staff_pending_requests')

@login_required
def staff_clearance_history(request):
    try:
        staff = request.user.staff
    except Staff.DoesNotExist:
        return redirect('login')

    clearance_requests = ClearanceRequest.objects.filter(
        office=staff.office,
        status__in=['approved', 'denied']
    ).select_related('student', 'student__user', 'student__course', 'reviewed_by').order_by('-reviewed_date')

    status = request.GET.get('status', '')
    if status:
        clearance_requests = clearance_requests.filter(status=status)

    school_year = request.GET.get('school_year', '')
    if school_year:
        clearance_requests = clearance_requests.filter(clearance__school_year=school_year)

    semester = request.GET.get('semester', '')
    if semester:
        clearance_requests = clearance_requests.filter(clearance__semester=semester)

    search_query = request.GET.get('search', '')
    if search_query:
        clearance_requests = clearance_requests.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    paginator = Paginator(clearance_requests, 10)
    page = request.GET.get('page')
    try:
        clearance_requests = paginator.page(page)
    except PageNotAnInteger:
        clearance_requests = paginator.page(1)
    except EmptyPage:
        clearance_requests = paginator.page(paginator.num_pages)

    return render(request, 'core/staff_clearance_history.html', {
        'clearance_requests': clearance_requests,
        'school_years': get_school_years(),
        'current_filters': {'status': status, 'school_year': school_year, 'semester': semester, 'search': search_query},
        'office': staff.office,
        'SEMESTER_CHOICES': SEMESTER_CHOICES,
    })

@login_required
def view_request(request, request_id):
    try:
        staff = request.user.staff
        request_obj = get_object_or_404(ClearanceRequest, id=request_id, office=staff.office)
        return render(request, 'core/staff_view_request.html', {'request_obj': request_obj})
    except Staff.DoesNotExist:
        messages.error(request, "Staff profile not found.")
        return redirect('home')

# Admin Views

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_dormitory_owners(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                # Get form data
                username = request.POST.get('username')
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                email = request.POST.get('email')
                password = request.POST.get('password')
                role = request.POST.get('role', 'Dormitory Owner')
                boarding_house_address = request.POST.get('boarding_house_address', '')

                # Check if username already exists
                if User.objects.filter(username=username).exists():
                    messages.error(request, f'Username "{username}" is already taken. Please choose another username.')
                    return redirect('admin_dormitory_owners')

                # Check if email already exists
                if User.objects.filter(email=email).exists():
                    messages.error(request, f'Email "{email}" is already in use. Please use another email address.')
                    return redirect('admin_dormitory_owners')

                # Get the dormitory office
                office = Office.objects.get(name="DORMITORY")

                # Create the user
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_active=True
                )

                # Create the dormitory owner
                staff = Staff.objects.create(
                    user=user,
                    office=office,
                    role=role,
                    is_dormitory_owner=True,
                    boarding_house_address=boarding_house_address
                )

                messages.success(request, f'Dormitory Owner {staff.get_full_name()} added successfully.')

            elif action == 'batch_assign':
                dormitory_owner_id = request.POST.get('dormitory_owner_id')
                student_ids = request.POST.getlist('student_ids')
                mark_as_boarders = request.POST.get('mark_as_boarders') == 'on'

                if not dormitory_owner_id or not student_ids:
                    messages.error(request, 'Missing required parameters for batch assignment.')
                    return redirect('admin_dormitory_owners')

                # Get the dormitory owner
                dormitory_owner = Staff.objects.get(id=dormitory_owner_id, is_dormitory_owner=True)

                # Assign students to the dormitory owner
                students = Student.objects.filter(id__in=student_ids)
                count = 0

                for student in students:
                    student.dormitory_owner = dormitory_owner
                    if mark_as_boarders:
                        student.is_boarder = True
                    student.save()
                    count += 1

                messages.success(request, f'{count} student(s) have been assigned to {dormitory_owner.get_full_name()}.')

            elif action == 'edit':
                staff = Staff.objects.get(id=request.POST.get('staff_id'))

                # Update role
                staff.role = request.POST.get('role')
                staff.save()

                messages.success(request, f'Dormitory Owner {staff.get_full_name()} updated successfully.')

            elif action == 'delete':
                staff = Staff.objects.get(id=request.POST.get('staff_id'))
                delete_type = request.POST.get('delete_type', 'safe')

                if delete_type == 'safe':
                    # Check if there are any students associated with this dormitory owner
                    if Student.objects.filter(dormitory_owner=staff).exists():
                        messages.error(
                            request,
                            f"Cannot safely delete dormitory owner '{staff.get_full_name()}' because they have associated students. "
                            f"Please reassign the students first or use Force Delete."
                        )
                    else:
                        # Safe to delete
                        staff_name = staff.get_full_name()
                        staff.delete()
                        messages.success(request, f'Dormitory Owner {staff_name} safely deleted.')

                elif delete_type == 'force':
                    # Force delete - will remove dormitory owner association from students
                    student_count = Student.objects.filter(dormitory_owner=staff).count()

                    # Use a transaction to ensure all operations succeed or fail together
                    from django.db import transaction
                    with transaction.atomic():
                        # Remove dormitory owner association from students
                        Student.objects.filter(dormitory_owner=staff).update(dormitory_owner=None, is_boarder=False)

                        # Delete the dormitory owner
                        staff_name = staff.get_full_name()
                        staff.delete()

                    messages.success(
                        request,
                        f'Dormitory Owner {staff_name} force deleted and {student_count} student(s) reassigned.'
                    )

        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        except Office.DoesNotExist:
            messages.error(request, 'Dormitory office not found. Please create an office named "DORMITORY" first.')
        except Staff.DoesNotExist:
            messages.error(request, 'Dormitory Owner not found.')
        except IntegrityError:
            messages.error(request, 'A user with this username or email already exists.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    # Get all dormitory owners
    dormitory_owners = Staff.objects.filter(is_dormitory_owner=True).select_related('user', 'office')

    # Add student count to each dormitory owner
    for owner in dormitory_owners:
        owner.student_count = Student.objects.filter(dormitory_owner=owner).count()

    return render(request, 'admin/dormitory_owners.html', {
        'dormitory_owners': dormitory_owners,
        'total_boarders': Student.objects.filter(is_boarder=True).count(),
        'total_dormitory_owners': dormitory_owners.count(),
        'unassigned_boarders': Student.objects.filter(is_boarder=True, dormitory_owner=None).count()
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_program_chairs(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                user = User.objects.get(id=request.POST.get('user'))
                dean = None
                if request.POST.get('dean'):
                    dean = Dean.objects.get(id=request.POST.get('dean'))

                # Check if user already has a program chair profile
                if hasattr(user, 'programchair'):
                    messages.error(request, f'User {user.get_full_name()} is already a program chair.')
                else:
                    # Create program chair
                    program_chair = ProgramChair.objects.create(
                        user=user,
                        dean=dean,
                        designation=request.POST.get('designation')
                    )

                    # Create staff entry if it doesn't exist
                    try:
                        office = Office.objects.get(name="PROGRAM CHAIR")
                        if not hasattr(user, 'staff'):
                            Staff.objects.create(user=user, office=office, role="Program Chair")
                    except Office.DoesNotExist:
                        messages.warning(request, 'PROGRAM CHAIR office not found. Please run migrations to create it.')

                    messages.success(request, f'Program Chair {program_chair.get_full_name()} added successfully.')

            elif action == 'edit':
                program_chair = ProgramChair.objects.get(id=request.POST.get('program_chair_id'))

                # Update dean
                if request.POST.get('dean'):
                    program_chair.dean = Dean.objects.get(id=request.POST.get('dean'))
                else:
                    program_chair.dean = None

                # Update designation
                program_chair.designation = request.POST.get('designation')
                program_chair.save()

                messages.success(request, f'Program Chair {program_chair.get_full_name()} updated successfully.')

            elif action == 'delete':
                program_chair = ProgramChair.objects.get(id=request.POST.get('program_chair_id'))
                delete_type = request.POST.get('delete_type', 'safe')

                if delete_type == 'safe':
                    # Check if there are any students associated with this program chair
                    if Student.objects.filter(program_chair=program_chair).exists():
                        messages.error(
                            request,
                            f"Cannot safely delete program chair '{program_chair.get_full_name()}' because they have associated students. "
                            f"Please reassign the students first or use Force Delete."
                        )
                    else:
                        # Safe to delete
                        program_chair_name = program_chair.get_full_name()
                        program_chair.delete()
                        messages.success(request, f'Program Chair {program_chair_name} safely deleted.')

                elif delete_type == 'force':
                    # Force delete - will remove program chair association from students
                    student_count = Student.objects.filter(program_chair=program_chair).count()

                    # Use a transaction to ensure all operations succeed or fail together
                    from django.db import transaction
                    with transaction.atomic():
                        # Remove program chair association from students
                        Student.objects.filter(program_chair=program_chair).update(program_chair=None)

                        # Delete the program chair
                        program_chair_name = program_chair.get_full_name()
                        program_chair.delete()

                    messages.success(
                        request,
                        f'Program Chair {program_chair_name} force deleted and {student_count} student(s) reassigned.'
                    )

        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        except Dean.DoesNotExist:
            messages.error(request, 'Dean not found.')
        except ProgramChair.DoesNotExist:
            messages.error(request, 'Program Chair not found.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    # Get all program chairs
    program_chairs = ProgramChair.objects.all().select_related('user', 'dean')

    # Add student count to each program chair
    for pc in program_chairs:
        pc.student_count = Student.objects.filter(program_chair=pc).count()

    # Get available users (those who are not already program chairs)
    program_chair_user_ids = program_chairs.values_list('user_id', flat=True)
    available_users = User.objects.exclude(id__in=program_chair_user_ids).filter(is_active=True)

    return render(request, 'admin/program_chairs.html', {
        'program_chairs': program_chairs,
        'available_users': available_users,
        'deans': Dean.objects.all()
    })
@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_reassign_students(request):
    selected_course_id = request.GET.get('course_to_delete') or None
    course = None
    students = None

    # If course_to_delete is provided in the URL, load that course
    if selected_course_id:
        try:
            course = Course.objects.get(id=selected_course_id)
            students = Student.objects.filter(course=course)
        except Course.DoesNotExist:
            messages.error(request, 'Course not found.')
            selected_course_id = None

    if request.method == 'POST':
        course_to_delete_id = request.POST.get('course_to_delete')
        course_to_reassign_id = request.POST.get('course_to_reassign')

        if course_to_delete_id and course_to_reassign_id:
            try:
                course_to_delete = Course.objects.get(id=course_to_delete_id)
                course_to_reassign = Course.objects.get(id=course_to_reassign_id)

                # Get all students in the course to delete
                students_to_reassign = Student.objects.filter(course=course_to_delete)
                count = students_to_reassign.count()

                if count > 0:
                    # Reassign all students to the new course
                    for student in students_to_reassign:
                        student.course = course_to_reassign
                        student.save()

                    # Now delete the course
                    course_to_delete.delete()

                    messages.success(
                        request,
                        f'Successfully reassigned {count} students from {course_to_delete.code} to {course_to_reassign.code} and deleted the course.'
                    )
                    return redirect('admin_courses')
                else:
                    # No students to reassign, just delete the course
                    course_to_delete.delete()
                    messages.success(request, f'Course {course_to_delete.code} deleted successfully.')
                    return redirect('admin_courses')

            except Course.DoesNotExist:
                messages.error(request, 'One or both courses not found.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        elif course_to_delete_id:
            # Just showing the students for the selected course
            selected_course_id = course_to_delete_id
            try:
                course = Course.objects.get(id=selected_course_id)
                students = Student.objects.filter(course=course)
            except Course.DoesNotExist:
                messages.error(request, 'Course not found.')

    # Get all courses with student count
    courses = Course.objects.annotate(student_count=Count('student')).all()

    return render(request, 'admin/reassign_students.html', {
        'courses': courses,
        'selected_course_id': selected_course_id,
        'course': course,
        'students': students,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    pending_approvals = Student.objects.filter(
        user__is_active=False
    ).select_related('user', 'course').order_by('-user__date_joined')

    return render(request, 'admin/dashboard.html', {
        'total_students': Student.objects.count(),
        'total_staff': Staff.objects.count(),
        'total_program_chairs': ProgramChair.objects.count(),
        'total_dormitory_owners': Staff.objects.filter(is_dormitory_owner=True).count(),
        'total_boarders': Student.objects.filter(is_boarder=True).count(),
        'clearance_stats': {
            'total': Clearance.objects.count(),
            'pending': ClearanceRequest.objects.filter(status='pending').count(),
            'approved': ClearanceRequest.objects.filter(status='approved').count(),
            'denied': ClearanceRequest.objects.filter(status='denied').count(),
        },
        'recent_clearances': Clearance.objects.select_related('student')[:5],
        'offices': Office.objects.annotate(
            staff_count=Count('staff'),
            pending_requests=Count('clearance_requests', filter=Q(clearance_requests__status='pending'))
        ),
        'pending_approvals': pending_approvals,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_profile(request):
    profile, created = UserProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST' and request.POST.get('action') == 'change_password':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        # Validate passwords
        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
        elif new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
        else:
            # Set new password with Django's built-in password hashing
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Password changed successfully. Please log in with your new password.')
            return redirect('login')

    return render(request, 'core/admin_profile.html', {
        'admin_user': request.user,
        'profile': profile
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_students(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        student_id = request.POST.get('student_id')

        if action == 'approve':
            try:
                student = Student.objects.get(id=student_id)
                student.is_approved = True
                student.approval_date = timezone.now()
                student.approval_admin = request.user
                student.user.is_active = True
                student.save()
                student.user.save()
                messages.success(request, f'Student {student.user.get_full_name()} approved successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')

        elif action == 'delete':
            try:
                student = Student.objects.get(id=student_id)
                student_name = student.user.get_full_name()
                # Delete the user which will cascade delete the student
                student.user.delete()
                messages.success(request, f'Student {student_name} deleted successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')

        elif action == 'add_student_form':
            return render(request, 'admin/add_student.html', {
                'courses': Course.objects.all(),
                'deans': Dean.objects.all().order_by('name'),
            })

        elif action == 'export_pdf':
            from core.pdf_utils import generate_students_pdf

            # Get the filtered students based on the current filters
            search_query = request.POST.get('search_query', '')
            course_filter = request.POST.get('course_filter', '')
            year_level_filter = request.POST.get('year_level_filter', '')
            status_filter = request.POST.get('status_filter', '')
            boarder_filter = request.POST.get('boarder_filter', '')

            # Start with all students
            students = Student.objects.select_related('user', 'course').all()

            # Apply filters
            if search_query:
                students = students.filter(
                    Q(user__first_name__icontains=search_query) |
                    Q(user__last_name__icontains=search_query) |
                    Q(student_id__icontains=search_query) |
                    Q(user__email__icontains=search_query)
                )

            if course_filter:
                students = students.filter(course_id=course_filter)

            if year_level_filter:
                students = students.filter(year_level=year_level_filter)

            if status_filter == 'active':
                students = students.filter(user__is_active=True)
            elif status_filter == 'inactive':
                students = students.filter(user__is_active=False)
            elif status_filter == 'pending':
                students = students.filter(user__is_active=False)

            if boarder_filter == '1':
                students = students.filter(is_boarder=True)
            elif boarder_filter == '0':
                students = students.filter(is_boarder=False)

            # Generate PDF
            pdf = generate_students_pdf(students, request)

            # Create the HttpResponse with PDF headers
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="students_report.pdf"'

            return response

        elif action == 'view_details':
            try:
                student = Student.objects.get(id=student_id)
                return render(request, 'admin/view_student.html', {
                    'student': student,
                })
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
                return redirect('admin_students')

        elif action == 'edit_form':
            try:
                student = Student.objects.get(id=student_id)
                # Get all dormitory owners for the dropdown
                dormitory_owners = Staff.objects.filter(is_dormitory_owner=True).select_related('user')

                return render(request, 'admin/edit_student.html', {
                    'student': student,
                    'courses': Course.objects.all(),
                    'dormitory_owners': dormitory_owners,
                })
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
                return redirect('admin_students')

        elif action == 'edit_student':
            try:
                student = Student.objects.get(id=student_id)
                user = student.user

                # Update user information
                user.first_name = request.POST.get('first_name')
                user.last_name = request.POST.get('last_name')
                user.email = request.POST.get('email')
                user.save()

                # Update student information
                student.contact_number = request.POST.get('contact_number')
                if request.POST.get('course'):
                    student.course_id = request.POST.get('course')
                if request.POST.get('year_level'):
                    student.year_level = request.POST.get('year_level')

                # Handle boarding information
                is_boarder = request.POST.get('is_boarder') == 'on'
                student.is_boarder = is_boarder

                # Handle dormitory owner assignment
                dormitory_owner_id = request.POST.get('dormitory_owner')
                if dormitory_owner_id:
                    student.dormitory_owner_id = dormitory_owner_id
                elif not is_boarder:
                    # If student is not a boarder, remove dormitory owner assignment
                    student.dormitory_owner = None

                # Handle boarder_since date
                boarder_since = request.POST.get('boarder_since')
                if is_boarder and boarder_since:
                    # Convert the datetime-local input to a Python datetime
                    try:
                        from datetime import datetime
                        student.boarder_since = datetime.strptime(boarder_since, '%Y-%m-%dT%H:%M')
                    except ValueError:
                        # If there's an error parsing the date, use current time
                        student.boarder_since = timezone.now()
                elif is_boarder and not student.boarder_since:
                    # If becoming a boarder and no date is set, use current time
                    student.boarder_since = timezone.now()

                student.is_approved = request.POST.get('is_approved') == 'on'
                student.user.is_active = student.is_approved
                student.save()
                user.save()

                messages.success(request, f'Student {student.user.get_full_name()} updated successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
            except Exception as e:
                messages.error(request, f'Error updating student: {str(e)}')

        elif action == 'add_student':
            try:
                # Create user
                user = User.objects.create_user(
                    username=request.POST.get('email').split('@')[0],  # Use first part of email as username
                    password='student123',  # Default password
                    email=request.POST.get('email'),
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    is_active=request.POST.get('auto_approve') == 'on'
                )

                # Generate student ID
                import random
                import string
                from django.utils.text import slugify

                year = timezone.now().year % 100
                first_chars = slugify(request.POST.get('first_name'))[:2].upper()
                last_chars = slugify(request.POST.get('last_name'))[:2].upper()
                random_digits = ''.join(random.choices(string.digits, k=4))
                student_id = f"{year}-{first_chars}{last_chars}-{random_digits}"

                # Get the course
                course_id = request.POST.get('course')

                # Get program chair if selected
                program_chair_id = request.POST.get('program_chair')
                program_chair = None
                if program_chair_id:
                    try:
                        program_chair = ProgramChair.objects.get(id=program_chair_id)
                    except ProgramChair.DoesNotExist:
                        pass

                # Create student
                student = Student.objects.create(
                    user=user,
                    student_id=student_id,
                    course_id=course_id,
                    program_chair=program_chair,
                    year_level=request.POST.get('year_level'),
                    contact_number=request.POST.get('contact_number'),
                    is_boarder=request.POST.get('is_boarder') == 'on',
                    is_approved=request.POST.get('auto_approve') == 'on'
                )

                # If no program chair was selected, try to assign one automatically
                if not program_chair:
                    course = Course.objects.get(id=course_id)
                    if course and course.dean:
                        # Try to find a program chair for this dean
                        auto_program_chair = ProgramChair.objects.filter(dean=course.dean).first()
                        if auto_program_chair:
                            student.program_chair = auto_program_chair
                            student.save()

                messages.success(request, f'Student {student.user.get_full_name()} added successfully with ID {student_id}.')
            except Exception as e:
                messages.error(request, f'Error adding student: {str(e)}')
                # Clean up if user was created but student wasn't
                if 'user' in locals() and 'student' not in locals():
                    user.delete()

    # Get query parameters for filtering
    search_query = request.GET.get('search', '')
    course_filter = request.GET.get('course', '')
    year_level_filter = request.GET.get('year_level', '')
    status_filter = request.GET.get('status', '')
    boarder_filter = request.GET.get('boarder', '')

    # Start with all students
    students = Student.objects.select_related('user', 'course').all()

    # Apply filters
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    if course_filter:
        students = students.filter(course_id=course_filter)

    if year_level_filter:
        students = students.filter(year_level=year_level_filter)

    if status_filter == 'active':
        students = students.filter(user__is_active=True)
    elif status_filter == 'inactive':
        students = students.filter(user__is_active=False)
    elif status_filter == 'pending':
        students = students.filter(is_approved=False)

    if boarder_filter == '1':
        students = students.filter(is_boarder=True)
    elif boarder_filter == '0':
        students = students.filter(is_boarder=False)

    return render(request, 'admin/students.html', {
        'students': students,
        'courses': Course.objects.all(),
        'search_query': search_query,
        'course_filter': course_filter,
        'year_level_filter': year_level_filter,
        'status_filter': status_filter,
        'boarder_filter': boarder_filter,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_staff(request):
    staff = Staff.objects.select_related('user', 'office').all().order_by('office__name', 'user__first_name')

    return render(request, 'admin/staff.html', {
        'staff': staff,
        'statistics': {
            'total_staff': staff.count(),
            'active_staff': staff.filter(user__is_active=True).count(),
            'total_offices': Office.objects.count(),
            'staff_by_office': Office.objects.annotate(staff_count=Count('staff')),
        }
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_staff_add(request):
    if request.method == 'POST':
        try:
            # Check if this will be a dormitory owner
            is_dormitory_owner = request.POST.get('is_dormitory_owner') == 'on'

            # Apply formatting to names
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            role = request.POST.get('role')

            # Apply capfirst to names (capitalize first letter, lowercase rest)
            first_name = first_name[0].upper() + first_name[1:].lower() if first_name else ''
            last_name = last_name[0].upper() + last_name[1:].lower() if last_name else ''
            # Apply uppercase to role
            role = role.upper() if role else ''

            user = User.objects.create_user(
                username=request.POST.get('username'),
                email=request.POST.get('email'),
                password=request.POST.get('password'),
                first_name=first_name,
                last_name=last_name,
                is_active=True
            )

            # Create staff object
            staff = Staff.objects.create(
                user=user,
                office_id=request.POST.get('office'),
                role=role,  # Use the uppercase role
                is_dormitory_owner=is_dormitory_owner
            )

            # Handle boarding house address if staff is a dormitory owner
            if is_dormitory_owner:
                staff.boarding_house_address = request.POST.get('boarding_house_address', '')
                staff.save()
            messages.success(request, 'Staff member added successfully.')
            return redirect('admin_staff')
        except Exception as e:
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error: {str(e)}')

    return render(request, 'admin/staff_add.html', {'offices': Office.objects.all()})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_staff_edit(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)

    if request.method == 'POST':
        try:
            # Check if this will be a dormitory owner
            is_dormitory_owner = request.POST.get('is_dormitory_owner') == 'on'

            # Staff is being updated to be a dormitory owner

            # Apply formatting to names
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            role = request.POST.get('role')

            # Apply capfirst to names (capitalize first letter, lowercase rest)
            first_name = first_name[0].upper() + first_name[1:].lower() if first_name else ''
            last_name = last_name[0].upper() + last_name[1:].lower() if last_name else ''
            # Apply uppercase to role
            role = role.upper() if role else ''

            staff.user.username = request.POST.get('username')
            staff.user.email = request.POST.get('email')
            staff.user.first_name = first_name
            staff.user.last_name = last_name
            if request.POST.get('password'):
                staff.user.set_password(request.POST.get('password'))
            staff.user.is_active = request.POST.get('is_active') == 'on'
            staff.user.save()

            staff.office_id = request.POST.get('office')
            staff.role = role
            staff.is_dormitory_owner = is_dormitory_owner

            # Handle boarding house address if staff is a dormitory owner
            if is_dormitory_owner:
                staff.boarding_house_address = request.POST.get('boarding_house_address', '')

            if 'profile_picture' in request.FILES:
                staff.profile_picture = request.FILES['profile_picture']
            staff.save()

            messages.success(request, 'Staff member updated successfully.')
            return redirect('admin_staff')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return render(request, 'admin/staff_edit.html', {
        'staff_member': staff,
        'offices': Office.objects.all(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_deans(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                dean = Dean.objects.create(
                    name=request.POST.get('name').upper(),
                    description=request.POST.get('description'),
                    logo=request.FILES.get('logo') if 'logo' in request.FILES else None
                )
                messages.success(request, f'Dean {dean.name} added.')
            elif action == 'edit':
                dean = Dean.objects.get(id=request.POST.get('dean_id'))
                dean.name = request.POST.get('name').upper()
                dean.description = request.POST.get('description')
                if 'logo' in request.FILES:
                    if dean.logo:
                        dean.logo.delete(save=False)
                    dean.logo = request.FILES['logo']
                dean.save()
                messages.success(request, f'Dean {dean.name} updated.')
            elif action == 'delete':
                dean = Dean.objects.get(id=request.POST.get('dean_id'))
                delete_type = request.POST.get('delete_type', 'safe')

                if delete_type == 'safe':
                    # Check if there are any courses associated with this dean
                    if dean.courses.exists():
                        messages.error(
                            request,
                            f"Cannot safely delete dean '{dean.name}' because it has associated courses. "
                            f"Please reassign the courses first or use a different delete option."
                        )
                    else:
                        # Safe to delete
                        dean_name = dean.name
                        if dean.logo:
                            dean.logo.delete(save=False)
                        dean.delete()
                        messages.success(request, f'Dean {dean_name} safely deleted.')

                elif delete_type == 'force':
                    # Force delete - will cascade delete all courses
                    course_count = dean.courses.count()
                    student_count = Student.objects.filter(course__dean=dean).count()

                    # If there are students, we need to handle them first
                    if student_count > 0:
                        messages.error(
                            request,
                            f"Cannot force delete dean '{dean.name}' because there are {student_count} students "
                            f"enrolled in its courses. Please reassign or delete these students first."
                        )
                    else:
                        # No students, safe to delete courses and dean
                        dean_name = dean.name
                        if dean.logo:
                            dean.logo.delete(save=False)

                        # Use a transaction to ensure all operations succeed or fail together
                        from django.db import transaction
                        with transaction.atomic():
                            # Delete all courses first
                            dean.courses.all().delete()
                            # Then delete the dean
                            dean.delete()

                        messages.success(
                            request,
                            f'Dean {dean_name} force deleted along with {course_count} course(s).'
                        )

                elif delete_type == 'cascade':
                    # Cascade delete - will delete everything including students
                    course_count = dean.courses.count()
                    student_count = Student.objects.filter(course__dean=dean).count()

                    # Use a transaction to ensure all operations succeed or fail together
                    from django.db import transaction
                    with transaction.atomic():
                        dean_name = dean.name
                        if dean.logo:
                            dean.logo.delete(save=False)

                        # Delete all students associated with this dean's courses
                        Student.objects.filter(course__dean=dean).delete()

                        # Delete all courses
                        dean.courses.all().delete()

                        # Delete the dean
                        dean.delete()

                    messages.success(
                        request,
                        f'Dean {dean_name} permanently deleted along with {course_count} course(s) and {student_count} student(s).'
                    )
            elif action == 'reassign':
                # Redirect to the reassign view
                dean_id = request.POST.get('dean_id')
                return redirect('admin_reassign_courses', dean_id=dean_id)
        except Dean.DoesNotExist:
            messages.error(request, 'Dean not found.')

    deans = Dean.objects.all()

    # Get additional statistics for the template
    total_courses = Course.objects.count()
    latest_dean = deans.order_by('-created_at').first() if deans.exists() else None

    # Add course count to each dean
    for dean in deans:
        dean.course_count = dean.courses.count()
        dean.student_count = Student.objects.filter(course__dean=dean).count()

    return render(request, 'admin/deans.html', {
        'deans': deans,
        'total_courses': total_courses,
        'latest_dean': latest_dean
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_reassign_courses(request, dean_id):
    """View to reassign courses from one dean to another"""
    try:
        source_dean = Dean.objects.get(id=dean_id)
    except Dean.DoesNotExist:
        messages.error(request, "Dean not found.")
        return redirect('admin_deans')

    # Get all courses for this dean
    courses = source_dean.courses.all()
    if not courses.exists():
        messages.warning(request, f"No courses found for dean '{source_dean.name}'.")
        return redirect('admin_deans')

    # Add student count to each course for display
    for course in courses:
        course.student_count = Student.objects.filter(course=course).count()

    # Get all other deans
    available_deans = Dean.objects.exclude(id=dean_id)

    if request.method == 'POST':
        target_dean_id = request.POST.get('target_dean')
        if not target_dean_id:
            messages.error(request, "Please select a target dean.")
            return redirect('admin_reassign_courses', dean_id=dean_id)

        try:
            target_dean = Dean.objects.get(id=target_dean_id)

            # Use a transaction to ensure all courses are reassigned or none
            from django.db import transaction
            with transaction.atomic():
                count = courses.update(dean=target_dean)

            messages.success(
                request,
                f"Successfully reassigned {count} courses from '{source_dean.name}' to '{target_dean.name}'."
            )
            return redirect('admin_deans')
        except Exception as e:
            messages.error(request, f"Error reassigning courses: {str(e)}")
            return redirect('admin_reassign_courses', dean_id=dean_id)

    return render(request, 'admin/reassign_dean_courses.html', {
        'source_dean': source_dean,
        'available_deans': available_deans,
        'courses': courses,
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_offices(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                office = Office.objects.create(
                    name=request.POST.get('name').upper(),
                    description=request.POST.get('description'),
                    location=request.POST.get('location'),
                    contact_number=request.POST.get('contact_number'),
                    email=request.POST.get('email')
                )
                messages.success(request, f'Office "{office.name}" added.')
            elif action == 'edit':
                office = Office.objects.get(id=request.POST.get('office_id'))
                office.name = request.POST.get('name').upper()
                office.description = request.POST.get('description')
                office.location = request.POST.get('location')
                office.contact_number = request.POST.get('contact_number')
                office.email = request.POST.get('email')
                office.save()
                messages.success(request, f'Office "{office.name}" updated.')
            elif action == 'delete':
                office = Office.objects.get(id=request.POST.get('office_id'))
                # Check if there are staff members associated with this office
                if office.staff.exists():
                    staff_list = ", ".join([staff.get_full_name() for staff in office.staff.all()[:5]])
                    if office.staff.count() > 5:
                        staff_list += f" and {office.staff.count() - 5} more"

                    messages.error(
                        request,
                        f"Cannot delete office '{office.name}' because it has associated staff members: {staff_list}. "
                        f"Please reassign or delete these staff members first."
                    )
                else:
                    # Safe to delete
                    office_name = office.name
                    office.delete()
                    messages.success(request, f'Office "{office_name}" deleted.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    offices = Office.objects.annotate(
        staff_count=Count('staff'),
        pending_requests=Count('clearance_requests', filter=Q(clearance_requests__status='pending')),
        completed_requests=Count('clearance_requests', filter=Q(clearance_requests__status='completed')),
        total_requests=Count('clearance_requests')
    ).order_by('name')

    return render(request, 'admin/offices.html', {
        'offices': offices,
        'total_offices': offices.count(),
        'total_staff': Staff.objects.count(),
        'total_pending_requests': sum(office.pending_requests for office in offices),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_courses(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'add':
                dean = get_object_or_404(Dean, id=request.POST.get('dean'))
                # Apply formatting to course code (uppercase)
                code = request.POST.get('code').upper()
                name = request.POST.get('name')

                Course.objects.create(
                    code=code,
                    name=name,
                    dean=dean
                )
                messages.success(request, f'Course {request.POST.get("code")} added.')
            elif action == 'delete':
                course = get_object_or_404(Course, id=request.POST.get('course_id'))

                # Check if there are students associated with this course
                students = Student.objects.filter(course=course)
                if students.exists():
                    student_list = ", ".join([f"{student.get_full_name()} ({student.student_id})" for student in students[:5]])
                    if students.count() > 5:
                        student_list += f" and {students.count() - 5} more"

                    error_message = f"Cannot delete course '{course.code}' because it is referenced by students: {student_list}. " \
                                   f"Please reassign these students to another course first."
                    messages.error(request, error_message)
                else:
                    course.delete()
                    messages.success(request, f'Course {course.code} deleted.')
            elif action == 'edit':
                course = get_object_or_404(Course, id=request.POST.get('course_id'))

                # Apply formatting to course code (uppercase)
                code = request.POST.get('code').upper()
                name = request.POST.get('name')

                course.code = code
                course.name = name
                course.dean = get_object_or_404(Dean, id=request.POST.get('dean'))
                course.save()
                messages.success(request, f'Course {course.code} updated.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('admin_courses')

    return render(request, 'admin/courses.html', {
        'courses': Course.objects.annotate(student_count=Count('student')).select_related('dean').all(),
        'deans': Dean.objects.all(),
        'total_students': Student.objects.count(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_clearances(request):
    clearances = Clearance.objects.select_related('student').all().order_by('-school_year', '-semester')
    return render(request, 'admin/clearances.html', {
        'clearances': clearances,
        'clearance_stats': {
            'pending': ClearanceRequest.objects.filter(status='PENDING').count(),
            'approved': ClearanceRequest.objects.filter(status='APPROVED').count(),
            'denied': ClearanceRequest.objects.filter(status='DENIED').count(),
        },
        'recent_clearances': Clearance.objects.select_related('student').order_by('-created_at')[:5],
    })
@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_pending_approvals(request):
    pending_students = Student.objects.filter(
        is_approved=False
    ).select_related('user', 'course')

    courses = Course.objects.all()

    return render(request, 'admin/pending_approvals.html', {
        'pending_students': pending_students,
        'total_pending': pending_students.count(),
        'courses': courses,
    })
@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_users(request):
    return render(request, 'admin/users.html', {
        'students': Student.objects.select_related('user', 'course').all(),
        'staff': Staff.objects.select_related('user', 'office').all(),
        'program_chairs': ProgramChair.objects.select_related('user').all(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_settings(request):
    # Get or create system settings
    from core.models import SystemSettings
    settings = SystemSettings.get_settings()

    if request.method == 'POST':
        # Determine which form was submitted
        form_type = request.POST.get('form_type')

        if form_type == 'academic_settings':
            # Handle academic settings updates
            school_year = request.POST.get('school_year')
            semester = request.POST.get('semester')

            # Validate school year format (YYYY-YYYY)
            try:
                year = int(school_year)
                settings.school_year = f"{year}-{year + 1}"
            except (ValueError, TypeError):
                messages.error(request, "Invalid school year format. Please enter a valid year.")

            # Validate semester
            if semester in dict(SEMESTER_CHOICES):
                settings.semester = semester
            else:
                messages.error(request, "Invalid semester selection.")

            # Save changes
            settings.updated_by = request.user
            settings.save()
            messages.success(request, "Academic settings updated successfully.")

        elif form_type == 'clearance_settings':
            # Handle clearance settings updates
            clearance_active = request.POST.get('clearance_active') == 'on'

            # Update settings
            settings.clearance_active = clearance_active
            settings.updated_by = request.user
            settings.save()

            status = "activated" if clearance_active else "deactivated"
            messages.success(request, f"Clearance system has been {status}.")

        elif form_type == 'system_settings':
            # Handle system settings updates
            maintenance_mode = request.POST.get('maintenance_mode') == 'on'
            email_notifications = request.POST.get('email_notifications') == 'on'

            # Update settings
            settings.maintenance_mode = maintenance_mode
            settings.email_notifications = email_notifications
            settings.updated_by = request.user
            settings.save()

            messages.success(request, "System settings updated successfully.")

    # Get semester choices for the template
    semester_choices = [(code, name) for code, name in SEMESTER_CHOICES]

    return render(request, 'admin/settings.html', {
        'settings': settings,
        'semester_choices': semester_choices,
        'current_time': timezone.now(),
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_reports(request):
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        school_year = request.POST.get('school_year')
        semester = request.POST.get('semester')
        format_type = request.POST.get('format_type', 'pdf')
        department = request.POST.get('department')

        if report_type == 'student_list':
            # Generate student list report
            students = Student.objects.select_related('user', 'course').all()

            # Filter by department (dean) if specified and not "all"
            if department and department != 'all' and department != '':
                students = students.filter(course__dean_id=department)

            # Check if there's data
            if not students.exists():
                messages.warning(request, "No student data found for the selected criteria.")
                return redirect('admin_reports')

            if format_type == 'pdf':
                from core.pdf_utils import generate_students_pdf
                pdf = generate_students_pdf(students, request, school_year, semester, department)
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="student_list_{school_year}_{semester}.pdf"'
                return response
            else:  # Excel format
                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="student_list_{school_year}_{semester}.xlsx"'

                wb = Workbook()
                ws = wb.active
                ws.title = "Student List"

                # Add headers
                headers = ['ID', 'Name', 'Course', 'Year Level', 'Contact Number']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Add data
                for row_num, student in enumerate(students, 2):
                    row_data = [
                        student.student_id,
                        f"{student.user.first_name} {student.user.last_name}",
                        student.course.name if student.course else 'Not assigned',
                        student.year_level,
                        student.contact_number or 'N/A'
                    ]

                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                wb.save(response)
                return response

        elif report_type == 'pending_clearances':
            # Generate pending clearances report
            clearance_requests = ClearanceRequest.objects.filter(
                school_year=school_year,
                semester=semester,
                status='pending'
            ).select_related('student', 'office')

            # Filter by department (dean) if specified and not "all"
            if department and department != 'all' and department != '':
                clearance_requests = clearance_requests.filter(student__course__dean_id=department)

            # Check if there's data
            if not clearance_requests.exists():
                messages.warning(request, "No pending clearance requests found for the selected criteria.")
                return redirect('admin_reports')

            if format_type == 'pdf':
                from core.pdf_utils import generate_pending_clearances_pdf
                pdf = generate_pending_clearances_pdf(clearance_requests, request, school_year, semester, department)
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="pending_clearances_{school_year}_{semester}.pdf"'
                return response
            else:  # Excel format
                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="pending_clearances_{school_year}_{semester}.xlsx"'

                wb = Workbook()
                ws = wb.active
                ws.title = "Pending Clearances"

                # Add headers
                headers = ['Student ID', 'Student Name', 'Office', 'Request Date']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Add data
                for row_num, request in enumerate(clearance_requests, 2):
                    row_data = [
                        request.student.student_id,
                        f"{request.student.user.first_name} {request.student.user.last_name}",
                        request.office.name,
                        request.request_date.strftime("%Y-%m-%d") if request.request_date else 'N/A'
                    ]

                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                wb.save(response)
                return response

        elif report_type == 'cleared_students':
            # Generate cleared students report
            clearances = Clearance.objects.filter(
                school_year=school_year,
                semester=semester,
                is_cleared=True
            ).select_related('student')

            # Filter by department (dean) if specified and not "all"
            if department and department != 'all' and department != '':
                clearances = clearances.filter(student__course__dean_id=department)

            # Check if there's data
            if not clearances.exists():
                messages.warning(request, "No cleared students found for the selected criteria.")
                return redirect('admin_reports')

            if format_type == 'pdf':
                from core.pdf_utils import generate_cleared_students_pdf
                pdf = generate_cleared_students_pdf(clearances, request, school_year, semester, department)
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="cleared_students_{school_year}_{semester}.pdf"'
                return response
            else:  # Excel format
                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="cleared_students_{school_year}_{semester}.xlsx"'

                wb = Workbook()
                ws = wb.active
                ws.title = "Cleared Students"

                # Add headers
                headers = ['Student ID', 'Student Name', 'Course', 'Year Level', 'Date Cleared']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Add data
                for row_num, clearance in enumerate(clearances, 2):
                    student = clearance.student
                    row_data = [
                        student.student_id,
                        f"{student.user.first_name} {student.user.last_name}",
                        student.course.name if student.course else 'Not assigned',
                        student.year_level,
                        clearance.cleared_date.strftime("%Y-%m-%d") if clearance.cleared_date else 'N/A'
                    ]

                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                wb.save(response)
                return response

        elif report_type == 'student_status':
            # Generate student status report
            students = Student.objects.select_related('user', 'course').all()

            # Filter by school year and semester if provided
            if school_year and semester:
                students = students.filter(clearances__school_year=school_year, clearances__semester=semester)

            if format_type == 'pdf':
                from core.pdf_utils import generate_students_pdf
                pdf = generate_students_pdf(students, request)
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="student_status_{school_year}_{semester}.pdf"'
                return response
            else:  # Excel format
                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="student_status_{school_year}_{semester}.xlsx"'

                wb = Workbook()
                ws = wb.active
                ws.title = "Student Status"

                # Add headers
                headers = ['ID', 'Name', 'Course', 'Year Level', 'Status', 'Contact Number']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Add data
                for row_num, student in enumerate(students, 2):
                    # Check if student has a clearance for the selected school year and semester
                    clearance = Clearance.objects.filter(
                        student=student,
                        school_year=school_year,
                        semester=semester
                    ).first() if school_year and semester else None

                    status = 'Cleared' if clearance and clearance.is_cleared else 'Pending'

                    row_data = [
                        student.student_id,
                        f"{student.user.first_name} {student.user.last_name}",
                        student.course.name if student.course else 'Not assigned',
                        student.year_level,
                        status,
                        student.contact_number or 'N/A'
                    ]

                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                wb.save(response)
                return response

        elif report_type == 'students_by_year_level':
            # Generate students by year level report
            students = Student.objects.select_related('user', 'course').all()

            # Filter by department (dean) if specified and not "all"
            if department and department != 'all' and department != '':
                students = students.filter(course__dean_id=department)

            # Check if there's data
            if not students.exists():
                messages.warning(request, "No student data found for the selected criteria.")
                return redirect('admin_reports')

            if format_type == 'pdf':
                from core.pdf_utils import generate_students_by_year_level_pdf
                pdf = generate_students_by_year_level_pdf(students, request, school_year, semester, department)
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="students_by_year_level_{school_year}_{semester}.pdf"'
                return response
            else:  # Excel format
                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="students_by_year_level_{school_year}_{semester}.xlsx"'

                wb = Workbook()
                ws = wb.active
                ws.title = "Students by Year Level"

                # Add title
                ws.merge_cells('A1:E1')
                title_cell = ws['A1']
                title_cell.value = f"Students by Year Level - {school_year} {semester}"
                title_cell.font = Font(size=16, bold=True)
                title_cell.alignment = Alignment(horizontal='center')

                # Group students by year level
                students_by_year = {}
                for student in students:
                    year_level = student.year_level
                    if year_level not in students_by_year:
                        students_by_year[year_level] = []
                    students_by_year[year_level].append(student)

                # Sort year levels
                sorted_years = sorted(students_by_year.keys())

                # Current row for writing data
                current_row = 3

                # Process each year level
                for year in sorted_years:
                    year_students = students_by_year[year]

                    # Year level header
                    year_text = {
                        1: "First Year Students",
                        2: "Second Year Students",
                        3: "Third Year Students",
                        4: "Fourth Year Students",
                        5: "Fifth Year Students"
                    }.get(year, f"Year {year} Students")

                    # Add year level header
                    ws.merge_cells(f'A{current_row}:E{current_row}')
                    header_cell = ws[f'A{current_row}']
                    header_cell.value = year_text
                    header_cell.font = Font(size=14, bold=True)
                    header_cell.alignment = Alignment(horizontal='center')
                    current_row += 1

                    # Add column headers
                    headers = ['Student ID', 'Name', 'Course', 'Contact', 'Status']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.value = header
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
                    current_row += 1

                    # Add student data
                    for student in year_students:
                        status = "Approved" if student.is_approved else "Pending"

                        row_data = [
                            student.student_id,
                            f"{student.user.first_name} {student.user.last_name}",
                            student.course.name if student.course else "N/A",
                            student.contact_number or "N/A",
                            status
                        ]

                        for col_num, cell_value in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = cell_value

                        current_row += 1

                    # Add year level summary
                    ws.merge_cells(f'A{current_row}:B{current_row}')
                    summary_cell = ws[f'A{current_row}']
                    summary_cell.value = f"Total {year_text}: {len(year_students)}"
                    summary_cell.font = Font(bold=True)
                    current_row += 2  # Add extra space between year levels

                # Add overall summary
                current_row += 1
                ws.merge_cells(f'A{current_row}:E{current_row}')
                overall_header = ws[f'A{current_row}']
                overall_header.value = "Overall Summary"
                overall_header.font = Font(size=14, bold=True)
                overall_header.alignment = Alignment(horizontal='center')
                current_row += 1

                # Add summary headers
                ws.cell(row=current_row, column=1).value = "Year Level"
                ws.cell(row=current_row, column=2).value = "Number of Students"
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2).font = Font(bold=True)
                current_row += 1

                # Add summary data
                total_students = 0
                for year in sorted_years:
                    year_text = {
                        1: "First Year",
                        2: "Second Year",
                        3: "Third Year",
                        4: "Fourth Year",
                        5: "Fifth Year"
                    }.get(year, f"Year {year}")

                    count = len(students_by_year[year])
                    total_students += count

                    ws.cell(row=current_row, column=1).value = year_text
                    ws.cell(row=current_row, column=2).value = count
                    current_row += 1

                # Add total row
                ws.cell(row=current_row, column=1).value = "Total"
                ws.cell(row=current_row, column=2).value = total_students
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2).font = Font(bold=True)

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(response)
                return response

        elif report_type == 'course_statistics':
            # Generate course statistics report
            courses = Course.objects.all()
            course_stats = []

            for course in courses:
                students = Student.objects.filter(course=course)
                total = students.count()

                if school_year and semester:
                    cleared = students.filter(
                        clearances__is_cleared=True,
                        clearances__school_year=school_year,
                        clearances__semester=semester
                    ).count()
                else:
                    cleared = 0

                pending = total - cleared
                clearance_rate = round((cleared / total * 100) if total > 0 else 0, 2)

                course_stats.append({
                    'course': course.name,
                    'total': total,
                    'cleared': cleared,
                    'pending': pending,
                    'clearance_rate': clearance_rate
                })

            if format_type == 'pdf':
                from core.pdf_utils import generate_course_statistics_pdf
                pdf = generate_course_statistics_pdf(course_stats, request, school_year, semester, department)
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="course_statistics_{school_year}_{semester}.pdf"'
                return response
            else:  # Excel format
                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="course_statistics_{school_year}_{semester}.xlsx"'

                wb = Workbook()
                ws = wb.active
                ws.title = "Course Statistics"

                # Add headers
                headers = ['Course', 'Total Students', 'Cleared', 'Pending', 'Clearance Rate']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Add data
                for row_num, stat in enumerate(course_stats, 2):
                    row_data = [
                        stat['course'],
                        stat['total'],
                        stat['cleared'],
                        stat['pending'],
                        f"{stat['clearance_rate']}%"
                    ]

                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                wb.save(response)
                return response

    return render(request, 'admin/reports.html', {
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
        'now': datetime.now(),
        'deans': Dean.objects.all(),
        'departments': Dean.objects.all(),  # Add departments for consistency
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def create_user(request):
    if request.method == 'POST':
        try:
            user = User.objects.create_user(
                username=request.POST.get('username'),
                password=request.POST.get('password'),
                email=request.POST.get('email'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name')
            )

            user_type = request.POST.get('user_type')
            if user_type == 'student':
                Student.objects.create(
                    user=user,
                    student_id=request.POST.get('student_id'),
                    course=Course.objects.get(id=request.POST.get('course')),
                    program_chair=ProgramChair.objects.get(id=request.POST.get('program_chair')),
                    year_level=request.POST.get('year_level'),
                    is_approved=True,
                    approval_date=timezone.now(),
                    approval_admin=request.user
                )
            elif user_type == 'staff':
                Staff.objects.create(
                    user=user,
                    office=Office.objects.get(id=request.POST.get('office')),
                    is_dormitory_owner=request.POST.get('is_dormitory_owner') == 'on'
                )
            elif user_type == 'program_chair':
                ProgramChair.objects.create(
                    user=user,
                    dean=Dean.objects.get(id=request.POST.get('dean'))
                )

            messages.success(request, f'User {user.username} created.')
            return redirect('admin_users')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return render(request, 'admin/create_user.html', {
        'courses': Course.objects.all(),
        'offices': Office.objects.all(),
        'program_chairs': ProgramChair.objects.all(),
        'deans': Dean.objects.all(),
    })

# Clearance Views
@login_required
def update_clearance_request(request, request_id):
    clearance_request = get_object_or_404(ClearanceRequest, pk=request_id)

    try:
        staff_member = request.user.staff

        if staff_member.office != clearance_request.office:
            messages.error(request, f"No permission for {clearance_request.office.name}")
            return redirect('staff_dashboard')

        if clearance_request.office.name == "DORMITORY":
            if not staff_member.is_dormitory_owner or clearance_request.student.dormitory_owner != staff_member:
                messages.error(request, "Dormitory clearance permission denied.")
                return redirect('staff_dashboard')

        if clearance_request.office.name.startswith('SSB'):
            if clearance_request.office.affiliated_dean != clearance_request.student.course.dean:
                messages.error(request, "SSB clearance permission denied.")
                return redirect('staff_dashboard')

    except Staff.DoesNotExist:
        messages.error(request, "Staff access required.")
        return redirect('login')

    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '')

        try:
            if action == 'approve':
                clearance_request.approve(staff_member)
                messages.success(request, f"Approved for {clearance_request.student.full_name}")
            elif action == 'deny':
                if not remarks:
                    messages.error(request, "Remarks required for denial.")
                    return redirect('staff_dashboard')
                clearance_request.deny(staff_member, remarks)
                messages.success(request, f"Denied for {clearance_request.student.full_name}")
            else:
                messages.error(request, "Invalid action.")
                return redirect('staff_dashboard')

            clearance = Clearance.objects.get(
                student=clearance_request.student,
                school_year=clearance_request.school_year,
                semester=clearance_request.semester
            )
            clearance.check_clearance()

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return redirect('staff_dashboard')

@login_required
def view_clearance_details(request, clearance_id):
    clearance = get_object_or_404(Clearance, id=clearance_id)

    if request.user.is_superuser:
        pass
    elif hasattr(request.user, 'programchair'):
        if clearance.student.course.dean != request.user.programchair.dean:
            messages.error(request, "Permission denied.")
            return redirect('program_chair_dashboard')
    elif hasattr(request.user, 'staff'):
        if not ClearanceRequest.objects.filter(clearance=clearance, office=request.user.staff.office).exists():
            messages.error(request, "Permission denied.")
            return redirect('staff_dashboard')
    elif hasattr(request.user, 'student'):
        if request.user.student != clearance.student:
            messages.error(request, "Permission denied.")
            return redirect('student_dashboard')
    else:
        messages.error(request, "Permission denied.")
        return redirect('home')

    clearance_requests = ClearanceRequest.objects.filter(clearance=clearance).select_related('office', 'reviewed_by')

    return render(request, 'core/clearance_details.html', {
        'clearance': clearance,
        'clearance_requests': clearance_requests,
        'user_type': ('admin' if request.user.is_superuser else
                     'program_chair' if hasattr(request.user, 'programchair') else
                     'staff' if hasattr(request.user, 'staff') else 'student')
    })

# Report Views
@login_required
def generate_reports(request):
    if request.method == 'POST':
        # Handle report generation
        pass

    return render(request, 'core/generate_reports.html', {
        'school_years': get_school_years(),
        'semesters': SEMESTER_CHOICES,
    })

@login_required
def check_clearance_status(request, student_id):
    """HTMX endpoint to check and update clearance status"""
    student = get_object_or_404(Student, id=student_id)

    # Get current school year and semester
    current_year = timezone.now().year
    school_year = f"{current_year}-{current_year + 1}"

    # Determine current semester based on month
    month = timezone.now().month
    if 1 <= month <= 5:  # January to May
        semester = "2ND"
    elif 6 <= month <= 10:  # June to October
        semester = "1ST"
    else:  # November to December
        semester = "SUM"

    # Get or create clearance
    clearance, created = Clearance.objects.get_or_create(
        student=student,
        school_year=school_year,
        semester=semester
    )

    # Check clearance status
    is_cleared = clearance.check_clearance()

    # Return just the clearance status cell for HTMX to swap
    return render(request, 'core/partials/clearance_status_cell.html', {
        'student': student,
        'clearance': clearance
    })

@login_required
def generate_report(request):
    if request.method == 'POST':
        school_year = request.POST.get('school_year')
        semester = request.POST.get('semester')
        report_type = request.POST.get('report_type')

        if report_type == 'pdf':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="clearance_report_{school_year}_{semester}.pdf"'
            # Add PDF generation logic
            return response
        elif report_type == 'excel':
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="clearance_report_{school_year}_{semester}.xlsx"'
            # Add Excel generation logic
            return response

        messages.success(request, 'Report generated successfully')
    return redirect('generate_reports')

@login_required
def export_clearances_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clearance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="44B78B", end_color="44B78B", fill_type="solid")

    headers = ['Student ID', 'Student Name', 'Course', 'Year Level', 'School Year', 'Semester',
              'Clearance Status', 'Date Cleared']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 15

    clearances = Clearance.objects.select_related('student', 'student__course').order_by(
        '-school_year', '-semester', 'student__student_id')

    for row, clearance in enumerate(clearances, 2):
        student = clearance.student
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.full_name)
        ws.cell(row=row, column=3, value=student.course.code)
        ws.cell(row=row, column=4, value=student.year_level)
        ws.cell(row=row, column=5, value=clearance.school_year)
        ws.cell(row=row, column=6, value=dict(SEMESTER_CHOICES)[clearance.semester])
        ws.cell(row=row, column=7, value="Cleared" if clearance.is_cleared else "Pending")
        ws.cell(row=row, column=8, value=clearance.cleared_date.strftime("%Y-%m-%d") if clearance.cleared_date else "N/A")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=clearance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

# Utility Views
@login_required
@require_POST
def update_profile_picture(request):
    try:
        if 'profile_picture' not in request.FILES:
            messages.error(request, 'No image provided')
            return redirect('home')

        if hasattr(request.user, 'student'):
            profile = request.user.student
            redirect_url = 'student_profile'
        elif hasattr(request.user, 'programchair'):
            profile = request.user.programchair
            redirect_url = 'program_chair_profile'
        elif hasattr(request.user, 'staff'):
            profile = request.user.staff
            redirect_url = 'staff_profile'
        elif request.user.is_superuser:
            profile, _ = UserProfile.objects.get_or_create(user=request.user)
            redirect_url = 'admin_profile'
        else:
            messages.error(request, 'Invalid user type')
            return redirect('home')

        if profile.profile_picture:
            profile.profile_picture.delete(save=False)
        profile.profile_picture = request.FILES['profile_picture']
        profile.save()

        messages.success(request, 'Profile picture updated')
        return redirect(redirect_url)
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('home')

@login_required
@require_POST
def update_contact_number(request):
    try:
        if not hasattr(request.user, 'student'):
            messages.error(request, 'Only students can update contact numbers')
            return redirect('home')

        student = request.user.student
        contact_number = request.POST.get('contact_number', '').strip()

        # Basic validation
        if contact_number and not contact_number.startswith('+'):
            contact_number = '+' + contact_number

        # Update the contact number
        student.contact_number = contact_number
        student.save()

        messages.success(request, 'Contact number updated successfully')
        return redirect('student_profile')
    except Exception as e:
        messages.error(request, f'Error updating contact number: {str(e)}')
        return redirect('student_profile')

@login_required
def print_permit(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    # Get current school year and semester from system settings
    from core.utils import get_current_school_year, get_current_semester
    school_year = get_current_school_year()
    semester = get_current_semester()

    # Get the clearance record
    clearance = Clearance.objects.filter(
        student=student,
        school_year=school_year,
        semester=semester
    ).first()

    # If no clearance exists, create one
    if not clearance:
        clearance = Clearance.objects.create(
            student=student,
            school_year=school_year,
            semester=semester
        )

    # Always check clearance status to ensure it's up to date
    clearance.check_clearance()

    return render(request, 'core/print_permit.html', {
        'student': student,
        'clearance': clearance,
        'logo_url': 'images/logo.png',
    })

@login_required
@user_passes_test(is_program_chair)
def batch_print_permits(request):
    """View to handle batch printing of permits"""
    # Get student IDs from session
    student_ids = request.session.get('batch_print_students', [])

    if not student_ids:
        messages.error(request, "No students selected for batch permit printing.")
        return redirect('program_chair_dashboard')

    # Get the students
    students = Student.objects.filter(id__in=student_ids)

    # Clear the session data
    if 'batch_print_students' in request.session:
        del request.session['batch_print_students']

    return render(request, 'core/batch_print_permits.html', {
        'students': students,
        'logo_url': 'images/logo.png',
    })

# API Endpoints
@login_required
def get_program_chairs(request, dean_id):
    program_chairs = ProgramChair.objects.filter(dean_id=dean_id).select_related('user')
    data = [{
        'id': pc.id,
        'user': {
            'full_name': f"{pc.user.get_full_name()}"
        }
    } for pc in program_chairs]
    return JsonResponse(data, safe=False)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_dormitory_owner_details(request, owner_id):
    try:
        owner = Staff.objects.get(id=owner_id, is_dormitory_owner=True)
        students = Student.objects.filter(dormitory_owner=owner).select_related('user')

        # Prepare owner data
        owner_data = {
            'id': owner.id,
            'full_name': owner.get_full_name(),
            'email': owner.user.email,
            'role': owner.role,
            'office': owner.office.name,
            'student_count': students.count(),
            'profile_picture': owner.get_profile_picture_url() if hasattr(owner, 'get_profile_picture_url') else None
        }

        # Prepare students data
        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'student_id': student.student_id,
                'full_name': student.get_full_name(),
                'course': student.course.name if student.course else 'Not assigned',
                'year_level': student.year_level
            })

        return JsonResponse({
            'success': True,
            'owner': owner_data,
            'students': students_data
        })
    except Staff.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Dormitory owner not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_unassigned_students(request):
    try:
        # Get students who are boarders but don't have a dormitory owner assigned
        students = Student.objects.filter(is_boarder=True, dormitory_owner=None).select_related('user', 'course')

        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'student_id': student.student_id,
                'full_name': student.get_full_name(),
                'course': student.course.name if student.course else 'Not assigned',
                'year_level': student.year_level
            })

        return JsonResponse({
            'success': True,
            'students': students_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_POST
def unassign_student(request):
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        dormitory_owner_id = data.get('dormitory_owner_id')

        if not student_id or not dormitory_owner_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing required parameters'
            }, status=400)

        # Get the student and verify they belong to the specified dormitory owner
        student = Student.objects.get(id=student_id, dormitory_owner_id=dormitory_owner_id)

        # Unassign the dormitory owner
        student.dormitory_owner = None
        student.save()

        return JsonResponse({
            'success': True,
            'message': f'Student {student.get_full_name()} has been unassigned from the dormitory owner.'
        })
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Student not found or not assigned to this dormitory owner'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_POST
def assign_students(request):
    try:
        dormitory_owner_id = request.POST.get('dormitory_owner_id')
        student_ids = request.POST.getlist('student_ids')
        mark_as_boarders = request.POST.get('mark_as_boarders') == 'on'

        if not dormitory_owner_id or not student_ids:
            return JsonResponse({
                'success': False,
                'message': 'Missing required parameters'
            }, status=400)

        # Get the dormitory owner
        dormitory_owner = Staff.objects.get(id=dormitory_owner_id, is_dormitory_owner=True)

        # Assign students to the dormitory owner
        students = Student.objects.filter(id__in=student_ids)
        count = 0
        skipped = 0
        already_assigned_students = []

        for student in students:
            # Check if student is already assigned to a different dormitory owner
            if student.dormitory_owner and student.dormitory_owner.id != dormitory_owner.id:
                skipped += 1
                already_assigned_students.append(f"{student.get_full_name()} ({student.student_id})")
                continue

            student.dormitory_owner = dormitory_owner
            if mark_as_boarders:
                student.is_boarder = True
                # Set boarder_since if it's not already set and student is becoming a boarder
                if mark_as_boarders and not student.boarder_since:
                    student.boarder_since = timezone.now()
            student.save()
            count += 1

        message = f'{count} student(s) have been assigned to {dormitory_owner.get_full_name()}.'
        if skipped > 0:
            message += f' {skipped} student(s) were skipped because they are already assigned to another dormitory owner.'

        return JsonResponse({
            'success': True,
            'message': message,
            'skipped_students': already_assigned_students if already_assigned_students else None
        })
    except Staff.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Dormitory owner not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_program_chair_details(request, program_chair_id):
    try:
        program_chair = ProgramChair.objects.get(id=program_chair_id)
        return JsonResponse({
            'id': program_chair.id,
            'user_id': program_chair.user.id,
            'full_name': program_chair.get_full_name(),
            'email': program_chair.user.email,
            'dean_id': program_chair.dean.id if program_chair.dean else None,
            'dean_name': program_chair.dean.name if program_chair.dean else None,
            'designation': program_chair.designation or '',
            'student_count': Student.objects.filter(program_chair=program_chair).count()
        })
    except ProgramChair.DoesNotExist:
        return JsonResponse({'error': 'Program Chair not found'}, status=404)

from django.http import JsonResponse
from django.core.exceptions import ObjectDoesNotExist
from .models import Course, Dean

def get_courses(request, dean_id):
    """
    Get courses for a specific dean. This endpoint is public and doesn't require authentication
    since it's used on the registration page.
    """
    try:
        # Debug logging
        print(f"Fetching courses for dean_id: {dean_id}")

        courses = Course.objects.filter(
            dean_id=dean_id,
            is_active=True
        ).order_by('code')

        # Debug logging
        print(f"Found courses: {[f'{c.code} (ID: {c.id})' for c in courses]}")

        data = [{
            'id': course.id,
            'code': course.code,
            'name': course.name
        } for course in courses]

        return JsonResponse(data, safe=False)

    except ObjectDoesNotExist:
        print(f"Dean not found with ID: {dean_id}")
        return JsonResponse(
            {'error': 'Dean not found'},
            status=404
        )
    except Exception as e:
        print(f"Error fetching courses: {str(e)}")
        return JsonResponse(
            {'error': str(e)},
            status=500
        )

@login_required
def get_offices(request, dean_id):
    offices = Office.objects.filter(affiliated_dean_id=dean_id)
    data = [{'id': office.id, 'name': office.name} for office in offices]
    return JsonResponse(data, safe=False)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_dean_details(request, dean_id):
    try:
        dean = Dean.objects.get(id=dean_id)
        return JsonResponse({
            'name': dean.name,
            'description': dean.description,
            'logo_url': dean.logo.url if dean.logo else None
        })
    except Dean.DoesNotExist:
        return JsonResponse({'error': 'Dean not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_user_details(request, user_id):
    try:
        student = Student.objects.select_related('user', 'course').get(user_id=user_id)
        return JsonResponse({
            'full_name': student.user.get_full_name(),
            'student_id': student.student_id,
            'course': student.course.name if student.course else 'N/A',
            'date_joined': student.user.date_joined.strftime('%B %d, %Y'),
            'email': student.user.email,
            'contact_number': student.contact_number or 'Not provided',
        })
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_student_details(request, student_id):
    try:
        student = Student.objects.select_related('user', 'course').get(id=student_id)
        return JsonResponse({
            'id': student.id,
            'full_name': f"{student.user.first_name} {student.user.last_name}",
            'first_name': student.user.first_name,
            'last_name': student.user.last_name,
            'student_id': student.student_id,
            'email': student.user.email,
            'contact_number': student.contact_number or '',
            'course': student.course.code if student.course else '',
            'course_id': student.course.id if student.course else '',
            'year_level': student.year_level,
            'date_applied': student.user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            'is_boarder': student.is_boarder,
            'is_approved': student.is_approved,
            'is_active': student.user.is_active,
            'profile_picture_url': student.get_profile_picture_url() if hasattr(student, 'get_profile_picture_url') else None,
            'created_at': student.created_at.strftime('%Y-%m-%d') if hasattr(student, 'created_at') and student.created_at else ''
        })
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_staff(request, staff_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        staff = Staff.objects.get(id=staff_id)
        user = staff.user
        staff.delete()
        user.delete()
        return JsonResponse({'success': True})
    except Staff.DoesNotExist:
        return JsonResponse({'error': 'Staff not found'}, status=404)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def approve_student(request, student_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        # Parse request body for notes
        data = json.loads(request.body)
        notes = data.get('notes', '')
        send_email_notification = data.get('send_email', True)  # Default to True

        student = Student.objects.get(id=student_id)
        student.approve_student(request.user)

        # If there are notes, you could store them in a new field or in a separate model
        # For now, we'll just log them
        if notes:
            logging.info(f"Approval notes for student {student.student_id}: {notes}")

            # You could add a field to store approval notes if needed:
            # student.approval_notes = notes
            # student.save()

        # Send email notification if requested
        email_status = None
        if send_email_notification:
            from core.email_utils import send_approval_email

            # Send the email
            success, result, details = send_approval_email(student, notes)

            if success:
                email_status = {
                    'sent': True,
                    'message': result,
                    'attempts': details.get('attempts', 1) if isinstance(details, dict) else 1,
                    'max_retries': getattr(settings, 'EMAIL_MAX_RETRIES', 3)
                }
                logging.info(f"Approval email sent to student {student.student_id} at {student.user.email} after {details.get('attempts', 1)} attempts")
            else:
                error_message = details.get('message', result) if isinstance(details, dict) else result
                email_status = {
                    'sent': False,
                    'error': error_message,
                    'error_type': details.get('error_type') if isinstance(details, dict) else 'UNKNOWN_ERROR',
                    'error_code': details.get('error_code') if isinstance(details, dict) else None,
                    'attempts': details.get('attempts', 1) if isinstance(details, dict) else 1,
                    'max_retries': getattr(settings, 'EMAIL_MAX_RETRIES', 3)
                }
                logging.warning(f"Failed to send approval email to student {student.student_id} after {details.get('attempts', 1)} attempts: {error_message}")

        return JsonResponse({'success': True, 'email_status': email_status})
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    except Exception as e:
        logging.error(f"Error approving student {student_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def reject_student(request, student_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        data = json.loads(request.body)
        reason = data.get('reason')
        send_email_notification = data.get('send_email', True)  # Default to True

        if not reason:
            return JsonResponse({'success': False, 'error': 'Reason is required'}, status=400)

        student = Student.objects.get(id=student_id)
        student.reject_student(reason)

        # Log the rejection
        logging.info(f"Student {student.student_id} rejected. Reason: {reason}")

        # Send email notification if requested
        email_status = None
        if send_email_notification:
            from core.email_utils import send_rejection_email

            # Send the email
            success, result, details = send_rejection_email(student, reason)

            if success:
                email_status = {
                    'sent': True,
                    'message': result,
                    'attempts': details.get('attempts', 1) if isinstance(details, dict) else 1,
                    'max_retries': getattr(settings, 'EMAIL_MAX_RETRIES', 3)
                }
                logging.info(f"Rejection email sent to student {student.student_id} at {student.user.email} after {details.get('attempts', 1)} attempts")
            else:
                error_message = details.get('message', result) if isinstance(details, dict) else result
                email_status = {
                    'sent': False,
                    'error': error_message,
                    'error_type': details.get('error_type') if isinstance(details, dict) else 'UNKNOWN_ERROR',
                    'error_code': details.get('error_code') if isinstance(details, dict) else None,
                    'attempts': details.get('attempts', 1) if isinstance(details, dict) else 1,
                    'max_retries': getattr(settings, 'EMAIL_MAX_RETRIES', 3)
                }
                logging.warning(f"Failed to send rejection email to student {student.student_id} after {details.get('attempts', 1)} attempts: {error_message}")

        return JsonResponse({'success': True, 'email_status': email_status})
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    except Exception as e:
        logging.error(f"Error rejecting student {student_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def office_detail_api(request, office_id):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    office = get_object_or_404(Office, id=office_id)
    data = {'name': office.name}
    if hasattr(office, 'description'):
        data['description'] = office.description
    return JsonResponse(data)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_approval_stats(request):
    total_pending = Student.objects.filter(user__is_active=False).count()
    return JsonResponse({'total_pending': total_pending})

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Course, Dean  # Import your models

@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_course_details(request, course_id):
    try:
        course = Course.objects.select_related('dean').get(id=course_id)
        return JsonResponse({
            'id': course.id,
            'code': course.code,
            'name': course.name,
            'dean_id': course.dean.id,
            'dean_name': course.dean.name
        })
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'}, status=404)

@require_http_methods(["GET"])
def get_courses_by_dean(request, dean_name):
    try:
        dean = Dean.objects.get(name=dean_name)
        courses = Course.objects.filter(dean=dean)
        data = [{'code': course.code, 'name': course.name} for course in courses]
        return JsonResponse(data, safe=False)
    except Dean.DoesNotExist:
        return JsonResponse([], safe=False)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def clearance_details(request, clearance_id):
    clearance = get_object_or_404(Clearance.objects.select_related(
        'student',
        'student__user',
        'student__course'
    ), id=clearance_id)

    clearance_requests = ClearanceRequest.objects.filter(
        clearance=clearance
    ).select_related('office', 'reviewed_by')

    return render(request, 'admin/clearance_details.html', {
        'clearance': clearance,
        'clearance_requests': clearance_requests,
    })











