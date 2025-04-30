import json
import io
from datetime import datetime

from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Count
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.template.loader import get_template

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from .models import Student, Clearance, ClearanceRequest, Office, Course, SEMESTER_CHOICES
from .utils import is_program_chair

@login_required
@user_passes_test(is_program_chair)
def student_clearance_details(request, student_id):
    """API endpoint to get clearance details for a specific student"""
    try:
        # Ensure the program chair has access to this student
        program_chair = request.user.programchair
        student = Student.objects.filter(
            id=student_id,
            course__dean=program_chair.dean
        ).select_related('user', 'course').first()

        if not student:
            return JsonResponse({'error': 'Student not found or access denied'}, status=404)

        # Get current school year and semester
        current_year = timezone.now().year
        school_year = request.GET.get('school_year', f"{current_year}-{current_year + 1}")
        semester = request.GET.get('semester', "1ST")

        # Get clearance for this student
        clearance = Clearance.objects.filter(
            student=student,
            school_year=school_year,
            semester=semester
        ).first()

        # Get office approvals
        office_approvals = []
        if clearance:
            clearance_requests = ClearanceRequest.objects.filter(
                clearance=clearance
            ).select_related('office', 'reviewed_by')

            for req in clearance_requests:
                office_approvals.append({
                    'office_id': req.office.id,
                    'office_name': req.office.name,
                    'is_approved': req.status == 'approved',
                    'status': req.status,
                    'remarks': req.remarks,
                    'reviewed_by': req.reviewed_by.user.get_full_name() if req.reviewed_by else None,
                    'reviewed_date': req.reviewed_date.strftime('%Y-%m-%d %H:%M') if req.reviewed_date else None
                })

        # Prepare response data
        data = {
            'id': student.id,
            'student_id': student.student_id,
            'full_name': student.user.get_full_name(),
            'email': student.user.email,
            'course': student.course.name,
            'course_code': student.course.code,
            'year_level': student.year_level,
            'clearance_details': {
                'id': clearance.id if clearance else None,
                'is_cleared': clearance.is_cleared if clearance else False,
                'cleared_date': clearance.cleared_date.strftime('%Y-%m-%d %H:%M') if clearance and clearance.cleared_date else None,
                'program_chair_approved': clearance.program_chair_approved if clearance else False,
                'school_year': school_year,
                'semester': semester,
            },
            'office_approvals': office_approvals
        }

        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(is_program_chair)
@require_http_methods(["POST"])
def approve_clearance(request):
    """API endpoint to approve a student's clearance"""
    try:
        data = request.POST if request.POST else json.loads(request.body)
        student_id = data.get('student_id')
        comment = data.get('comment', '')

        # Ensure the program chair has access to this student
        program_chair = request.user.programchair
        student = Student.objects.filter(
            id=student_id,
            course__dean=program_chair.dean
        ).first()

        if not student:
            return JsonResponse({'success': False, 'error': 'Student not found or access denied'}, status=404)

        # Get current school year and semester
        current_year = timezone.now().year
        school_year = data.get('school_year', f"{current_year}-{current_year + 1}")
        semester = data.get('semester', "1ST")

        # Get clearance for this student
        clearance = Clearance.objects.filter(
            student=student,
            school_year=school_year,
            semester=semester
        ).first()

        if not clearance:
            return JsonResponse({'success': False, 'error': 'Clearance not found'}, status=404)

        if not clearance.is_cleared:
            return JsonResponse({'success': False, 'error': 'Cannot approve clearance that is not cleared by all offices'}, status=400)

        # Approve the clearance
        clearance.program_chair_approved = True
        clearance.save()

        # TODO: Create an activity log entry if needed

        return JsonResponse({'success': True, 'message': 'Clearance approved successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(is_program_chair)
@require_http_methods(["POST"])
def deny_clearance(request):
    """API endpoint to deny a student's clearance"""
    try:
        data = request.POST if request.POST else json.loads(request.body)
        student_id = data.get('student_id')
        comment = data.get('comment', '')

        if not comment:
            return JsonResponse({'success': False, 'error': 'Comment is required when denying a clearance'}, status=400)

        # Ensure the program chair has access to this student
        program_chair = request.user.programchair
        student = Student.objects.filter(
            id=student_id,
            course__dean=program_chair.dean
        ).first()

        if not student:
            return JsonResponse({'success': False, 'error': 'Student not found or access denied'}, status=404)

        # Get current school year and semester
        current_year = timezone.now().year
        school_year = data.get('school_year', f"{current_year}-{current_year + 1}")
        semester = data.get('semester', "1ST")

        # Get clearance for this student
        clearance = Clearance.objects.filter(
            student=student,
            school_year=school_year,
            semester=semester
        ).first()

        if not clearance:
            return JsonResponse({'success': False, 'error': 'Clearance not found'}, status=404)

        # Deny the clearance by setting it to not cleared
        clearance.is_cleared = False
        clearance.program_chair_approved = False
        clearance.save()

        # Create a new clearance request for the program chair office
        program_chair_office, _ = Office.objects.get_or_create(
            name="Program Chair Office",
            defaults={"description": "Final clearance approval"}
        )

        clearance_request, _ = ClearanceRequest.objects.get_or_create(
            student=student,
            clearance=clearance,
            office=program_chair_office,
            school_year=school_year,
            semester=semester,
            defaults={
                'status': 'denied',
                'remarks': comment,
                'reviewed_by': None,  # We don't have a Staff object for program chair
                'reviewed_date': timezone.now()
            }
        )

        if clearance_request.status != 'denied':
            clearance_request.status = 'denied'
            clearance_request.remarks = comment
            clearance_request.reviewed_date = timezone.now()
            clearance_request.save()

        # TODO: Create an activity log entry if needed

        return JsonResponse({'success': True, 'message': 'Clearance denied successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(is_program_chair)
def students_batch_details(request):
    """API endpoint to get details for a batch of students"""
    try:
        # Handle both form data and JSON data
        if request.method == 'POST':
            if request.content_type == 'application/json':
                try:
                    data = json.loads(request.body)
                except json.JSONDecodeError:
                    return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
            else:
                data = request.POST
        else:
            return JsonResponse({'success': False, 'error': 'Only POST requests are supported'}, status=405)

        student_ids = data.get('student_ids', [])

        if not student_ids:
            return JsonResponse({'success': False, 'error': 'No student IDs provided'}, status=400)

        # Ensure the program chair has access to these students
        program_chair = request.user.programchair
        students = Student.objects.filter(
            id__in=student_ids,
            course__dean=program_chair.dean
        ).select_related('user', 'course')

        # Get current school year and semester
        current_year = timezone.now().year
        school_year = data.get('school_year', f"{current_year}-{current_year + 1}")
        semester = data.get('semester', "1ST")

        # Prepare student data
        student_data = []
        for student in students:
            clearance = Clearance.objects.filter(
                student=student,
                school_year=school_year,
                semester=semester
            ).first()

            status = "Not Started"
            if clearance:
                status = "Cleared" if clearance.is_cleared else "Pending"

            # Get profile picture URL safely
            profile_picture_url = None
            if hasattr(student, 'get_profile_picture_url'):
                try:
                    profile_picture_url = student.get_profile_picture_url()
                except:
                    profile_picture_url = '/static/img/default-profile.png'

            student_data.append({
                'id': student.id,
                'full_name': student.user.get_full_name(),
                'student_id': student.student_id,
                'course': student.course.code,
                'year_level': student.year_level,
                'status': status,
                'profile_picture': profile_picture_url
            })

        return JsonResponse({'success': True, 'students': student_data})
    except Exception as e:
        import traceback
        print(traceback.format_exc())  # Log the full error for debugging
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(is_program_chair)
def batch_approval_details(request, batch_id):
    """API endpoint to get details for a batch approval"""
    try:
        # This is a placeholder since we don't have a BatchApproval model
        # In a real implementation, you would fetch the batch approval details from the database

        # Generate some sample data based on the batch_id
        # In a real implementation, this would come from the database
        if batch_id == 1:
            # Sample data for a recent batch approval
            approval_date = timezone.now() - timezone.timedelta(days=1)
            approval_type = 'final'
            comments = 'End of semester clearance approval'
            student_count = 15
        else:
            # Sample data for an older batch approval
            approval_date = timezone.now() - timezone.timedelta(days=3)
            approval_type = 'permit'
            comments = 'Permit issuance for graduating students'
            student_count = 8

        # Get some actual students to show in the details
        program_chair = request.user.programchair
        students = Student.objects.filter(
            course__dean=program_chair.dean
        ).select_related('user', 'course')[:student_count]

        student_data = []
        for student in students:
            student_data.append({
                'id': student.id,
                'full_name': student.user.get_full_name(),
                'student_id': student.student_id,
                'course': student.course.code,
                'year_level': student.year_level,
                'profile_picture': student.get_profile_picture_url() if hasattr(student, 'get_profile_picture_url') else None
            })

        return JsonResponse({
            'id': batch_id,
            'date': approval_date.strftime('%Y-%m-%d %H:%M'),
            'approved_by': request.user.get_full_name(),
            'type': approval_type.title(),
            'school_year': f"{timezone.now().year}-{timezone.now().year + 1}",
            'semester': '1ST',
            'comments': comments,
            'students': student_data
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())  # Log the full error for debugging
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(is_program_chair)
def program_chair_reports(request, report_type):
    """API endpoint to get report data for program chair"""
    try:
        program_chair = request.user.programchair

        # Get current school year and semester
        current_year = timezone.now().year
        school_year = request.GET.get('school_year', f"{current_year}-{current_year + 1}")
        semester = request.GET.get('semester', "1ST_MID")
        year_level = request.GET.get('year_level')
        student_status = request.GET.get('student_status', 'all')  # all, cleared, pending

        # Get all students under this program chair's dean
        students = Student.objects.filter(course__dean=program_chair.dean)

        # Filter by year level if provided
        if year_level and year_level.isdigit():
            students = students.filter(year_level=int(year_level))

        # Get clearances for the selected school year and semester
        clearances = Clearance.objects.filter(
            student__in=students,
            school_year=school_year,
            semester=semester
        )

        # Filter by student status if provided
        if student_status == 'cleared':
            clearances = clearances.filter(is_cleared=True)
            students = students.filter(id__in=clearances.values_list('student_id', flat=True))
        elif student_status == 'pending':
            clearances = clearances.filter(is_cleared=False)
            students = students.filter(id__in=clearances.values_list('student_id', flat=True))

        if report_type == 'year_level':
            # Generate report data by year level
            data = []
            for year in range(1, 5):  # 1st to 4th year
                year_students = students.filter(year_level=year)
                total = year_students.count()

                if total == 0:
                    continue

                year_clearances = clearances.filter(student__in=year_students)
                cleared = year_clearances.filter(is_cleared=True).count()
                pending = year_clearances.filter(is_cleared=False).count()
                not_started = total - (cleared + pending)

                data.append({
                    'year_level': f"{year}{'st' if year == 1 else 'nd' if year == 2 else 'rd' if year == 3 else 'th'} Year",
                    'total': total,
                    'cleared': cleared,
                    'pending': pending,
                    'not_started': not_started
                })

            return JsonResponse({'success': True, 'data': data})

        elif report_type == 'course':
            # Generate report data by course
            data = []
            courses = Course.objects.filter(dean=program_chair.dean)

            for course in courses:
                course_students = students.filter(course=course)
                total = course_students.count()

                if total == 0:
                    continue

                course_clearances = clearances.filter(student__in=course_students)
                cleared = course_clearances.filter(is_cleared=True).count()
                pending = course_clearances.filter(is_cleared=False).count()

                data.append({
                    'course_code': course.code,
                    'course_name': course.name,
                    'total': total,
                    'cleared': cleared,
                    'pending': pending
                })

            return JsonResponse({'success': True, 'data': data})

        elif report_type == 'office':
            # Generate report data by office
            data = []
            offices = Office.objects.all()

            for office in offices:
                # Get clearance requests for this office
                requests = ClearanceRequest.objects.filter(
                    clearance__in=clearances,
                    office=office
                )

                total = requests.count()

                if total == 0:
                    continue

                approved = requests.filter(status='approved').count()
                pending = requests.filter(status='pending').count()

                # Calculate average processing time
                approved_requests = requests.filter(
                    status='approved',
                    reviewed_date__isnull=False
                )

                total_days = 0
                for req in approved_requests:
                    if req.reviewed_date and req.request_date:
                        days = (req.reviewed_date - req.request_date).days
                        total_days += max(0, days)  # Ensure non-negative

                avg_days = total_days / approved_requests.count() if approved_requests.count() > 0 else 0

                data.append({
                    'office_name': office.name,
                    'total': total,
                    'approved': approved,
                    'pending': pending,
                    'average_days': avg_days
                })

            return JsonResponse({'success': True, 'data': data})

        elif report_type == 'timeline':
            # Generate timeline report data
            # This would typically show clearance progress over time
            # For simplicity, we'll generate sample data
            data = []

            # Get the start date (beginning of semester)
            start_date = timezone.now().replace(month=1, day=1) if semester == '1ST' else \
                        timezone.now().replace(month=6, day=1) if semester == '2ND' else \
                        timezone.now().replace(month=4, day=1)  # Summer

            # Generate weekly data for the past 6 weeks
            for i in range(6):
                week_date = start_date + timezone.timedelta(weeks=i)

                # In a real implementation, you would query the database for actual data
                # Here we're generating sample data
                completion_rate = min(100, i * 15 + 10)  # Increases each week

                data.append({
                    'date': week_date.strftime('%Y-%m-%d'),
                    'new_clearances': 50 - i * 5,  # Decreases each week
                    'completed_clearances': 10 + i * 8,  # Increases each week
                    'permits_issued': 5 + i * 7,  # Increases each week
                    'completion_rate': completion_rate
                })

            return JsonResponse({'success': True, 'data': data})

        else:
            return JsonResponse({'success': False, 'error': 'Invalid report type'}, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(is_program_chair)
def export_program_chair_report(request, format_type):
    """API endpoint to export report data for program chair"""
    try:
        program_chair = request.user.programchair

        # Get parameters
        school_year = request.GET.get('school_year')
        semester = request.GET.get('semester')
        report_type = request.GET.get('report_type')
        year_level = request.GET.get('year_level')
        student_status = request.GET.get('student_status', 'all')  # all, cleared, pending

        if not all([school_year, semester, report_type]):
            return JsonResponse({'success': False, 'error': 'Missing required parameters'}, status=400)

        # Get all students under this program chair's dean
        students = Student.objects.filter(course__dean=program_chair.dean)

        # Filter by year level if provided
        if year_level and year_level.isdigit():
            students = students.filter(year_level=int(year_level))

        # Get clearances for the selected school year and semester
        clearances = Clearance.objects.filter(
            student__in=students,
            school_year=school_year,
            semester=semester
        )

        # Filter by student status if provided
        if student_status == 'cleared':
            clearances = clearances.filter(is_cleared=True)
            students = students.filter(id__in=clearances.values_list('student_id', flat=True))
        elif student_status == 'pending':
            clearances = clearances.filter(is_cleared=False)
            students = students.filter(id__in=clearances.values_list('student_id', flat=True))

        # Get semester display name for the title
        semester_display = dict(SEMESTER_CHOICES).get(semester, semester)

        # Get report data based on report type
        if report_type == 'year_level':
            data = get_year_level_report_data(students, clearances)
            title = f"Student's List - SY: {school_year} {semester_display} Records"
            headers = ['Year Level', 'Total Students', 'Cleared', 'Pending', 'Not Started', 'Clearance Rate']

        elif report_type == 'course':
            data = get_course_report_data(students, clearances, program_chair)
            title = f"Clearance Status by Course - {school_year}, {semester} Semester"
            headers = ['Course Code', 'Course Name', 'Total Students', 'Cleared', 'Pending', 'Clearance Rate']

        elif report_type == 'office':
            data = get_office_report_data(clearances)
            title = f"Clearance Status by Office - {school_year}, {semester} Semester"
            headers = ['Office', 'Total Requests', 'Approved', 'Pending', 'Avg. Processing Time', 'Approval Rate']

        elif report_type == 'timeline':
            data = get_timeline_report_data(semester, school_year, students, clearances)
            title = f"Clearance Timeline - {school_year}, {semester} Semester"
            headers = ['Date', 'New Clearances', 'Completed Clearances', 'Permits Issued', 'Completion Rate']

        else:
            return JsonResponse({'success': False, 'error': 'Invalid report type'}, status=400)

        # Special case for cleared students report
        if report_type == 'cleared_students':
            # Always get only cleared students, regardless of the student_status filter
            # This ensures we're always getting cleared students for this specific report
            cleared_clearances = Clearance.objects.filter(
                student__in=students,
                school_year=school_year,
                semester=semester,
                is_cleared=True
            ).select_related('student', 'student__user', 'student__course')

            # Apply year level filter if provided
            if year_level and year_level.isdigit():
                cleared_clearances = cleared_clearances.filter(student__year_level=int(year_level))

            # Log the count for debugging
            print(f"Found {cleared_clearances.count()} cleared students for export")

            if format_type == 'pdf':
                from core.pdf_utils import generate_cleared_students_pdf
                pdf = generate_cleared_students_pdf(cleared_clearances, request, school_year, semester)
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="cleared_students_{school_year}_{semester}.pdf"'
                return response
            elif format_type == 'excel':
                return export_cleared_students_excel(cleared_clearances, school_year, semester)

        # Export based on format type for other report types
        if format_type == 'pdf':
            return export_as_pdf(data, title, headers)
        elif format_type == 'excel':
            return export_as_excel(data, title, headers, report_type)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid export format'}, status=400)

    except Exception as e:
        import traceback
        print(traceback.format_exc())  # Log the full error for debugging
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# Helper functions for report data
def get_year_level_report_data(students, clearances):
    data = []
    for year in range(1, 5):  # 1st to 4th year
        year_students = students.filter(year_level=year)
        total = year_students.count()

        if total == 0:
            continue

        year_clearances = clearances.filter(student__in=year_students)
        cleared = year_clearances.filter(is_cleared=True).count()
        pending = year_clearances.filter(is_cleared=False).count()
        not_started = total - (cleared + pending)
        clearance_rate = round((cleared / total) * 100, 1) if total > 0 else 0

        year_suffix = 'st' if year == 1 else 'nd' if year == 2 else 'rd' if year == 3 else 'th'
        data.append([
            f"{year}{year_suffix} Year",
            total,
            cleared,
            pending,
            not_started,
            f"{clearance_rate}%"
        ])
    return data

def get_course_report_data(students, clearances, program_chair):
    data = []
    courses = Course.objects.filter(dean=program_chair.dean)

    for course in courses:
        course_students = students.filter(course=course)
        total = course_students.count()

        if total == 0:
            continue

        course_clearances = clearances.filter(student__in=course_students)
        cleared = course_clearances.filter(is_cleared=True).count()
        pending = course_clearances.filter(is_cleared=False).count()
        clearance_rate = round((cleared / total) * 100, 1) if total > 0 else 0

        data.append([
            course.code,
            course.name,
            total,
            cleared,
            pending,
            f"{clearance_rate}%"
        ])
    return data

def get_office_report_data(clearances):
    data = []
    offices = Office.objects.all()

    for office in offices:
        # Get clearance requests for this office
        requests = ClearanceRequest.objects.filter(
            clearance__in=clearances,
            office=office
        )

        total = requests.count()

        if total == 0:
            continue

        approved = requests.filter(status='approved').count()
        pending = requests.filter(status='pending').count()

        # Calculate average processing time
        approved_requests = requests.filter(
            status='approved',
            reviewed_date__isnull=False
        )

        total_days = 0
        for req in approved_requests:
            if req.reviewed_date and req.request_date:
                days = (req.reviewed_date - req.request_date).days
                total_days += max(0, days)  # Ensure non-negative

        avg_days = round(total_days / approved_requests.count(), 1) if approved_requests.count() > 0 else 0
        approval_rate = round((approved / total) * 100, 1) if total > 0 else 0

        data.append([
            office.name,
            total,
            approved,
            pending,
            f"{avg_days} days",
            f"{approval_rate}%"
        ])
    return data

def get_timeline_report_data(semester, school_year, students, clearances):
    data = []

    # Parse the school year to get the start year
    try:
        start_year = int(school_year.split('-')[0])
    except (ValueError, IndexError):
        start_year = timezone.now().year

    # Determine semester start date based on semester and school year
    from datetime import datetime, timezone as dt_timezone
    if semester.startswith('1ST'):
        # First semester typically starts in August/September
        start_date = datetime(start_year, 8, 1).replace(tzinfo=dt_timezone.utc)
    elif semester.startswith('2ND'):
        # Second semester typically starts in January/February
        start_date = datetime(start_year + 1, 1, 1).replace(tzinfo=dt_timezone.utc)
    else:  # Summer
        # Summer typically starts in May/June
        start_date = datetime(start_year + 1, 5, 1).replace(tzinfo=dt_timezone.utc)

    # Generate weekly data for 6 weeks
    total_students = students.count()

    for i in range(6):
        week_date = start_date + timezone.timedelta(weeks=i)
        week_end = week_date + timezone.timedelta(weeks=1)

        # For real implementation, we would query actual data from the database
        # Here we're using a combination of real data and simulated progression

        # Calculate how many clearances were created by this week (simulated)
        new_clearances = int(total_students * min(1.0, (i + 1) / 6))

        # Calculate completed clearances with a realistic progression
        completed_factor = min(1.0, (i + 1) / 8)  # Slower than creation rate
        completed_clearances = int(total_students * completed_factor * 0.8)  # Assume 80% completion rate

        # Calculate permits issued (typically follows completed clearances with a delay)
        permits_factor = min(1.0, (i + 0.5) / 8)  # Slightly behind completion rate
        permits_issued = int(total_students * permits_factor * 0.7)  # Assume 70% get permits

        # Calculate completion rate
        completion_rate = round((completed_clearances / total_students) * 100 if total_students > 0 else 0, 1)

        data.append([
            week_date.strftime('%Y-%m-%d'),
            new_clearances,
            completed_clearances,
            permits_issued,
            f"{completion_rate}%"
        ])
    return data

# Function to export cleared students as Excel
def export_cleared_students_excel(clearances, school_year, semester):
    """Export cleared students data to Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="cleared_students_{school_year}_{semester}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleared Students"

    # Add title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Cleared Students - {school_year}, {semester}"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Add date
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.font = Font(size=10)
    date_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Student ID', 'Student Name', 'Course', 'Year Level', 'Contact Number', 'Date Cleared']
    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for row_num, clearance in enumerate(clearances, header_row + 1):
        student = clearance.student
        year_level_text = {
            1: "1st Year",
            2: "2nd Year",
            3: "3rd Year",
            4: "4th Year",
            5: "5th Year"
        }.get(student.year_level, f"{student.year_level} Year")

        cleared_date = clearance.cleared_date.strftime("%Y-%m-%d") if clearance.cleared_date else "N/A"

        row_data = [
            student.student_id,
            f"{student.user.first_name} {student.user.last_name}",
            student.course.name if student.course else "N/A",
            year_level_text,
            student.contact_number or "N/A",
            cleared_date
        ]

        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center' if col_num > 1 else 'left')

            # Add alternating row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook to the response
    wb.save(response)
    return response

# Export functions
def export_as_pdf(data, title, headers):
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()

    # Create the PDF object, using the buffer as its "file"
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )

    # Create styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=12
    )

    # Create the content elements
    elements = []

    # Add title
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.25*inch))

    # Add date
    date_text = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 0.25*inch))

    # Add table with data
    table_data = [headers] + data
    table = Table(table_data, repeatRows=1)

    # Style the table
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ])

    # Add alternating row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)

    table.setStyle(table_style)
    elements.append(table)

    # Build the PDF
    doc.build(elements)

    # Get the value of the BytesIO buffer
    pdf = buffer.getvalue()
    buffer.close()

    # Create the HttpResponse with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.pdf"'
    response.write(pdf)

    return response

def export_as_excel(data, title, headers, report_type):
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()

    # Add title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Add date
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.font = Font(size=10)
    date_cell.alignment = Alignment(horizontal='center')

    # Add headers
    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for row_num, row_data in enumerate(data, header_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center' if col_num > 1 else 'left')

            # Add alternating row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Auto-adjust column widths
    column_widths = {}

    # Start from row 4 (header row) to avoid merged cells
    for row in ws.iter_rows(min_row=header_row):
        for cell in row:
            if not isinstance(cell, MergedCell):  # Skip merged cells
                column = cell.column_letter
                if cell.value:
                    # Calculate the length of the cell value
                    cell_length = len(str(cell.value))
                    # Update the maximum width for this column
                    if column not in column_widths or cell_length > column_widths[column]:
                        column_widths[column] = cell_length

    # Apply the calculated widths
    for column, width in column_widths.items():
        adjusted_width = (width + 2) * 1.2  # Add some padding
        ws.column_dimensions[column].width = adjusted_width

    # Create the HttpResponse with Excel headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response
